# ui.py — Review & Ticksheet (titles visible, exports always accessible)
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from tkinter import font as tkfont
import pathlib, subprocess, sys, threading, csv, time, re, datetime, sqlite3
from feature_dictionary import FeatureDictionary

CHECKED = "☑"
UNCHECKED = "☐"

ROOT = pathlib.Path(__file__).parent.resolve()
MASTERLIST_DIR = ROOT / "masterlists"
SESSIONS_DIR   = ROOT / "sessions"
EXPORTS_DIR    = ROOT / "exports"

# Row shape (UI): [nr, name, lv_match, en_match, final_match, confidence, include_sym, lv_text, en_text, consensus_reasoning]
# Fixed column mapping: confidence removed, columns properly aligned
UI_COLUMNS = [
    "Nr", "Variable Name",
    "LV Match", "EN Match", "Final Match", "Include",
    "LV Text", "EN Text", "Consensus Reasoning"
]
PIPELINE_COLUMNS = ["Nr Code","Variable Name","Is TT","Match (Yes/No/Maybe)","Matching Text","LLM_Reason"]
DUAL_PIPELINE_COLUMNS = ["Nr Code","Variable Name","Is TT","LV_Match","EN_Match","Final_Match","LV_Text","EN_Text","LV_Reason","EN_Reason"]
# Sequential MEGA has: Nr Code, Variable Name, Is TT, LV Match, LV Text, EN Match, EN Text, Consensus Vote, Consensus Text, Consensus Reasoning, Include
SEQUENTIAL_MEGA_COLUMNS = ["Nr Code","Variable Name","Is TT","LV Match","LV Text","EN Match","EN Text","Consensus Vote","Consensus Text","Consensus Reasoning","Include"]

def list_models():
    """Get available models, excluding EN versions"""
    return [p.stem for p in MASTERLIST_DIR.glob("*.csv") if not p.stem.endswith("_en")]
def model_exports_dir(model: str) -> pathlib.Path:
    p = EXPORTS_DIR / model
    p.mkdir(parents=True, exist_ok=True)
    return p
def working_path(model: str) -> pathlib.Path: return model_exports_dir(model) / "ticksheet_working.csv"

def find_latest_ticksheet(model: str):
    folder = model_exports_dir(model)
    files = list(folder.glob("ticksheet_*.csv")) + list(folder.glob("ticksheet_reviewed_*.csv"))
    files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return files[0] if files else None

def read_csv_any(path: pathlib.Path):
    with open(path, encoding="utf-8", newline="") as f:
        rdr = csv.reader(f)
        rows = list(rdr)
    header = rows[0] if rows else []
    data = rows[1:] if header else rows
    return header, data

def write_csv(path: pathlib.Path, header, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        if header: w.writerow(header)
        w.writerows(rows)

def _trim140(s: str) -> str:
    if s is None: return ""
    s = s.strip()
    return s if len(s) <= 140 else (s[:137] + "…")

def _letters_present(s: str) -> bool:
    # Unicode-safe check for any letter
    return any(ch.isalpha() for ch in (s or ""))

_STOPWORDS = {
    # tiny mix of EN/LV function words to ignore for token overlap
    "un","ar","vai","bez","pie","par","no","uz","in","on","and","or","with","w","the","a","an","of","to","for","at","by"
}

def _tokens4(text: str):
    text = (text or "").lower()
    # split on non-letters/digits (keep diacritics via str.isalnum)
    raw = []
    buf = []
    for ch in text:
        if ch.isalnum():
            buf.append(ch)
        else:
            if buf:
                raw.append("".join(buf))
                buf=[]
    if buf: raw.append("".join(buf))
    return {t for t in raw if len(t) >= 4 and t not in _STOPWORDS}

def evidence_sanity_check(match: str, snippet: str, variable_name: str):
    """
    Returns (match, snippet) possibly downgraded if evidence is weak or off-topic.
    Rules:
    - If snippet has no letters or length < 2 → blank snippet
    If match==Yes → downgrade to Maybe.
    - If match==Yes but snippet shares no >=4-letter token with Variable Name → downgrade to Maybe.
    - Trim snippet to ≤140 chars.
    """
    match = (match or "").strip().capitalize() if match else ""
    snippet = _trim140(snippet or "")
    if snippet and (not _letters_present(snippet) or len(snippet) < 2):
        return ("Maybe" if match == "Yes" else match), ""
    if match == "Yes" and snippet:
        vn_tokens = _tokens4(variable_name)
        snip_tokens = _tokens4(snippet)
        if vn_tokens and snip_tokens and vn_tokens.isdisjoint(snip_tokens):
            # no concept overlap → probably a wrong numeric/size-led match
            return "Maybe", snippet
    return match, snippet

class ReviewTable(ttk.Treeview):
    def __init__(self, master):
        super().__init__(master, columns=UI_COLUMNS, show="headings", selectmode="browse")
        for col in UI_COLUMNS:
            self.heading(col, text=col)
            self.column(col, anchor="w", stretch=True, width=100)

        # Configure column widths for optimized 9-column display
        self.column("Nr", width=35, stretch=False)
        self.column("Variable Name", width=280, stretch=True)
        self.column("LV Match", width=80, stretch=False)
        self.column("EN Match", width=80, stretch=False)
        self.column("Final Match", width=100, stretch=False)
        self.column("Include", width=60, stretch=False)
        self.column("LV Text", width=200, stretch=True)      # Latvian evidence text
        self.column("EN Text", width=200, stretch=True)      # English evidence text
        self.column("Consensus Reasoning", width=250, stretch=True)  # Detailed reasoning

        # Modern table styling with contemporary design
        style = ttk.Style()
        style.theme_use('clam')  # Use clam theme as base

        # Modern header styling with professional appearance
        style.configure("Treeview.Heading",
                       relief="flat",
                       borderwidth=0,
                       background="#F8FAFC",  # Light gray background
                       foreground="#1E293B",  # Dark slate text
                       font=("Segoe UI", 11, "bold"),  # Increased from 10 to 11
                       padding=(10, 8))

        # Modern table body styling with clean design
        style.configure("Treeview",
                       background="#FFFFFF",  # Pure white background
                       foreground="#1E293B",  # Dark slate text
                       fieldbackground="#FFFFFF",
                       borderwidth=0,
                       relief="flat",
                       font=("Segoe UI", 10),  # Increased from 9 to 10
                       rowheight=34)  # Increased from 32 to 34 for better readability

        # Modern selection and interaction styling
        style.map("Treeview",
                 background=[('selected', '#1E3A8A')],  # Professional blue
                 foreground=[('selected', '#FFFFFF')])

        # Enhanced row styling with subtle alternating colors
        self.tag_configure("evenrow", background="#FFFFFF")  # Pure white
        self.tag_configure("oddrow", background="#F8FAFC")   # Very light gray

        # colors for normal rows only
        self.tag_configure("YES", background="#d9f2d9")
        self.tag_configure("NO", background="#f8d7da")
        self.tag_configure("MAYBE", background="#fff3cd")

        # titles: smaller, bold font - not oversized
        base_font = tkfont.nametofont("TkDefaultFont")
        self.title_font = tkfont.Font(family=base_font.cget("family"),
                                      size=base_font.cget("size")+1,  # Only +1 instead of +2
                                      weight="bold")
        self.tag_configure("TT", background="#F0F9FF", font=self.title_font, foreground="#1E40AF")

        self.rows = []
        self.bind("<Button-1>", self.on_single_click)
        self.bind("<Double-1>", self.on_double_click)

    def _to_symbol(self, yn: str) -> str:
        return CHECKED if str(yn).strip().upper() in ("Y","YES","1","TRUE", CHECKED) else UNCHECKED

    def _tags_for(self, row):
        # Check if this is a title row by looking for "—" in the name
        if len(row) > 1 and "—" in str(row[1]):
            return ["TT"]
        # For regular rows, use Final Match column (index 4 in new 10-column structure)
        final_match = row[4] if len(row) > 4 else ""
        if final_match == "Yes": return ["YES"]
        if final_match == "No": return ["NO"]
        if final_match == "Maybe": return ["MAYBE"]
        return []

    def load_from_pipeline_rows(self, rows6):
        """Load single-language pipeline results (old format)"""
        self.delete(*self.get_children())
        self.rows = []
        for r in rows6:
            r = (r + [""]*6)[:6]
            nr, name, is_tt, match, snippet, reason = r
            if (is_tt or "").upper() == "Y":
                # titles: loud & clean, no checkbox, no color
                name = f"— {name.upper()} —"
                # Convert to new 10-column format: [nr, name, lv_match, en_match, final_match, confidence, include, lv_text, en_text, consensus_reasoning]
                row = [nr, name, "", "", "", "", "", "", "", ""]
            else:
                match, snippet = evidence_sanity_check(match, snippet, name)
                include_sym = self._to_symbol("Y" if match in ("Yes","Maybe") else "N")
                # Convert to new 10-column format: [nr, name, lv_match, en_match, final_match, confidence, include, lv_text, en_text, consensus_reasoning]
                row = [nr, name, "", "", match, "", include_sym, snippet, "", reason]
            self.rows.append(row)
            self.insert("", "end", iid=str(len(self.rows)-1), values=row, tags=self._tags_for(row))

    def load_from_reviewed_rows(self, rows7):
        """Load reviewed results (old format with Include column)"""
        self.delete(*self.get_children())
        self.rows = []
        for r in rows7:
            r = (r + [""]*7)[:7]
            nr, name, is_tt, match, include_raw, snippet, reason = r
            if (is_tt or "").upper() == "Y":
                name = f"— {name.upper()} —"
                # Convert to new 10-column format: [nr, name, lv_match, en_match, final_match, confidence, include, lv_text, en_text, consensus_reasoning]
                row = [nr, name, "", "", "", "", "", "", "", ""]
            else:
                match, snippet = evidence_sanity_check(match, snippet, name)
                include_sym = self._to_symbol(include_raw)
                # Convert to new 10-column format: [nr, name, lv_match, en_match, final_match, confidence, include, lv_text, en_text, consensus_reasoning]
                row = [nr, name, "", "", match, "", include_sym, snippet, "", reason]
            self.rows.append(row)
            self.insert("", "end", iid=str(len(self.rows)-1), values=row, tags=self._tags_for(row))

    def load_from_dual_results(self, lv_rows, en_rows, final_rows):
        """Load dual-language results with separate columns for LV and EN"""
        self.delete(*self.get_children())
        self.rows = []

        for lv_r, en_r, final_r in zip(lv_rows, en_rows, final_rows):
            lv_r = (lv_r + [""]*6)[:6]
            en_r = (en_r + [""]*6)[:6]
            final_r = (final_r + [""]*6)[:6]

            nr, name, is_tt = final_r[0], final_r[1], final_r[2]

            if (is_tt or "").upper() == "Y":
                # Section headers
                name = f"— {name.upper()} —"
                # Convert to new 10-column format: [nr, name, lv_match, en_match, final_match, confidence, include, lv_text, en_text, consensus_reasoning]
                row = [nr, name, "", "", "", "", "", "", "", ""]
            else:
                # Regular features
                lv_match, lv_snippet, lv_reason = lv_r[3], lv_r[4], lv_r[5]
                en_match, en_snippet, en_reason = en_r[3], en_r[4], en_r[5]
                final_match, final_snippet, final_reason = final_r[3], final_r[4], final_r[5]

                # Clean up reasons to remove language prefixes for cleaner display
                lv_reason = lv_reason.replace("LV: ", "") if lv_reason else ""
                en_reason = en_reason.replace("EN: ", "") if en_reason else ""

                include_sym = self._to_symbol("Y" if final_match in ("Yes","Maybe") else "N")

                # New 10-column format: [nr, name, lv_match, en_match, final_match, confidence, include, lv_text, en_text, consensus_reasoning]
                row = [nr, name, lv_match, en_match, final_match, "", include_sym, lv_snippet, en_snippet, final_reason]

            self.rows.append(row)
            self.insert("", "end", iid=str(len(self.rows)-1), values=row, tags=self._tags_for(row))

    def load_from_dual_csv(self, dual_rows):
        """Load from actual Learning Sequential MEGA CSV format (11 columns)"""
        self.delete(*self.get_children())
        self.rows = []

        # Detect format based on header or first data row
        if not dual_rows:
            return

        # Check if first row is header
        first_row = dual_rows[0] if dual_rows else []
        is_header = any(col.lower() in ('nr code', 'variable name', 'nr', 'number') for col in (first_row[:2] if first_row else []))

        data_rows = dual_rows[1:] if is_header else dual_rows

        for row in data_rows:
            row = list(row)  # Ensure it's a list

            # Handle actual Learning Sequential MEGA format (11 columns)
            # Actual format: Nr Code, Variable Name, Is TT, LV Match, EN Match, Final Match, Include, LV Text, EN Text, Consensus Reasoning, Accepted
            if len(row) >= 11:
                row = (row + [""]*11)[:11]
                nr, name, is_tt, lv_match, en_match, final_match, include, lv_text, en_text, consensus_reasoning, accepted = row

                if (is_tt or "").upper() == "Y":
                    # Section headers
                    name = f"— {name.upper()} —"
                    ui_row = [nr, name, "", "", "", "", "", "", ""]
                else:
                    # Regular features - Fix the data mapping issues
                    # Convert include to symbol if it's not already
                    if include in (CHECKED, UNCHECKED):
                        include_sym = include
                    else:
                        include_sym = CHECKED if str(include).strip().upper() in ("Y","YES","1","TRUE", CHECKED) else UNCHECKED

                    # Fix the data mapping issues:
                    # 1. EN Text often contains vote info like "Yes (3Y/0M/0N)" - extract the vote part for Final Match
                    # 2. EN Match often contains the actual evidence text
                    # 3. LV Text contains LV evidence

                    # Clean final match - use Final Match column if valid, otherwise extract from EN Text
                    final_match_clean = final_match if final_match in ["Yes", "No", "Maybe"] else ""
                    if not final_match_clean and en_text:
                        # EN Text might contain vote like "Yes (3Y/0M/0N)" - extract the result
                        if "(" in en_text and ")" in en_text:
                            vote_part = en_text.split("(")[0].strip()
                            if vote_part in ["Yes", "No", "Maybe"]:
                                final_match_clean = vote_part

                    # Fix evidence text mapping:
                    # LV Text should contain Latvian evidence
                    lv_text_clean = (lv_text or "").strip().strip('"')

                    # EN Text should contain English evidence, but if it contains vote info, use EN Match instead
                    if en_text and ("(" in en_text and ")" in en_text):
                        # EN Text contains vote info, use EN Match for evidence
                        en_text_clean = (en_match or "").strip().strip('"')
                    else:
                        # EN Text is actual evidence
                        en_text_clean = (en_text or "").strip().strip('"')

                    # Clean consensus reasoning
                    consensus_reasoning_clean = (consensus_reasoning or "").strip().strip('"')

                    # Ensure matches are proper Yes/No/Maybe format
                    lv_match_clean = lv_match if lv_match in ["Yes", "No", "Maybe"] else ""

                    # For EN Match, if it contains evidence text, convert to Yes/No based on whether there's evidence
                    if en_match and en_match not in ["Yes", "No", "Maybe"]:
                        en_match_clean = "Yes" if en_match.strip() else "No"
                    else:
                        en_match_clean = en_match if en_match in ["Yes", "No", "Maybe"] else ""

                    # New 9-column UI format: [nr, name, lv_match, en_match, final_match, include, lv_text, en_text, consensus_reasoning]
                    ui_row = [nr, name, lv_match_clean, en_match_clean, final_match_clean, include_sym, lv_text_clean, en_text_clean, consensus_reasoning_clean]

            else:
                # Legacy format - pad to 9 columns
                row = (row + [""]*9)[:9]
                ui_row = row

            self.rows.append(ui_row)
            self.insert("", "end", iid=str(len(self.rows)-1), values=ui_row, tags=self._tags_for(ui_row))

    def current_rows_ui(self): return [r[:] for r in self.rows]
    def set_rows_ui(self, ui_rows):
        self.delete(*self.get_children())
        self.rows = [r[:] for r in ui_rows]
        for i, r in enumerate(self.rows):
            self.insert("", "end", iid=str(i), values=r, tags=self._tags_for(r))

    def on_single_click(self, event):
        region = self.identify("region", event.x, event.y)
        if region != "cell": return
        row_id = self.identify_row(event.y)
        col_id = self.identify_column(event.x)
        if not row_id or not col_id: return
        idx = int(row_id)
        col_idx = int(col_id.replace("#","")) - 1
        INCLUDE_COL_INDEX = 5  # Fixed: Include column is now at index 5 in new 9-column structure
        if col_idx != INCLUDE_COL_INDEX: return
        if "—" in str(self.rows[idx][1]): return  # titles untickable (check by name format)
        self.rows[idx][5] = CHECKED if self.rows[idx][5] == UNCHECKED else UNCHECKED  # Include column index 5
        self.item(row_id, values=self.rows[idx], tags=self._tags_for(self.rows[idx]))

    def on_double_click(self, event):
        region = self.identify("region", event.x, event.y)
        if region != "cell": return
        row_id = self.identify_row(event.y)
        col_id = self.identify_column(event.x)
        if not row_id or not col_id: return
        idx = int(row_id)
        col_idx = int(col_id.replace("#","")) - 1
        if "—" in str(self.rows[idx][1]): return  # titles read-only

        # Column indices for new 10-column structure:
        # 0=Nr, 1=Variable Name, 2=LV Match, 3=EN Match, 4=Final Match, 5=Confidence, 6=Include, 7=LV Text, 8=EN Text, 9=Consensus Reasoning

        if col_idx == 2:  # LV Match
            choice = simpledialog.askstring("Edit LV Match", "Enter Match (Yes/No/Maybe):", initialvalue=self.rows[idx][2])
            if not choice: return
            choice = choice.strip().capitalize()
            if choice not in {"Yes","No","Maybe"}:
                return messagebox.showwarning("Invalid", "Use exactly: Yes, No, or Maybe.")
            self.rows[idx][2] = choice
        elif col_idx == 3:  # EN Match
            choice = simpledialog.askstring("Edit EN Match", "Enter Match (Yes/No/Maybe):", initialvalue=self.rows[idx][3])
            if not choice: return
            choice = choice.strip().capitalize()
            if choice not in {"Yes","No","Maybe"}:
                return messagebox.showwarning("Invalid", "Use exactly: Yes, No, or Maybe.")
            self.rows[idx][3] = choice
        elif col_idx == 4:  # Final Match
            choice = simpledialog.askstring("Edit Final Match", "Enter Match (Yes/No/Maybe):", initialvalue=self.rows[idx][4])
            if not choice: return
            choice = choice.strip().capitalize()
            if choice not in {"Yes","No","Maybe"}:
                return messagebox.showwarning("Invalid", "Use exactly: Yes, No, or Maybe.")
            self.rows[idx][4] = choice
            # auto-toggle include based on match
            if self.rows[idx][6] in (CHECKED, UNCHECKED):  # Include is now index 6
                self.rows[idx][6] = CHECKED if choice in ("Yes","Maybe") else UNCHECKED
        elif col_idx == 7:  # LV Text
            new_val = simpledialog.askstring("Edit LV Text", "LV Evidence Text:", initialvalue=self.rows[idx][7])
            if new_val is None: return
            new_val = _trim140(new_val)
            self.rows[idx][7] = new_val
        elif col_idx == 8:  # EN Text
            new_val = simpledialog.askstring("Edit EN Text", "EN Evidence Text:", initialvalue=self.rows[idx][8])
            if new_val is None: return
            new_val = _trim140(new_val)
            self.rows[idx][8] = new_val
        elif col_idx == 9:  # Consensus Reasoning
            label = "Consensus Reasoning"
            new_val = simpledialog.askstring(f"Edit {label}", f"{label}:", initialvalue=self.rows[idx][9])
            if new_val is None: return
            self.rows[idx][9] = new_val
        else:
            return
        self.item(row_id, values=self.rows[idx], tags=self._tags_for(self.rows[idx]))

    def tick_all(self):
        for i in range(len(self.rows)):
            if "—" not in str(self.rows[i][1]):  # Not a title row
                self.rows[i][6] = CHECKED  # Include is now index 6
                self.item(str(i), values=self.rows[i], tags=self._tags_for(self.rows[i]))
    def untick_all(self):
        for i in range(len(self.rows)):
            if "—" not in str(self.rows[i][1]):  # Not a title row
                self.rows[i][6] = UNCHECKED  # Include is now index 6
                self.item(str(i), values=self.rows[i], tags=self._tags_for(self.rows[i]))

    def _rows_for_export_with_accepted(self):
        rows = []
        for row in self.rows:
            # Ensure consistent handling for new 9-column format
            row = (row + [""]*9)[:9]  # Pad to 9 columns

            # New format: [nr, name, lv_match, en_match, final_match, include_sym, lv_text, en_text, consensus_reasoning]
            nr, name, lv_match, en_match, final_match, include_sym, lv_text, en_text, consensus_reasoning = row[:9]

            if "—" in str(name):  # Title row
                yn = ""
                accepted = ""
            else:
                yn = "Y" if include_sym == CHECKED else "N"
                accepted = "X" if include_sym == CHECKED else ""

            # Export format matching legacy structure
            is_tt = "Y" if "—" in str(name) else "N"
            matching_text = lv_text or en_text or ""  # Use LV text primarily, fallback to EN
            llm_reason = consensus_reasoning or ""

            rows.append([nr, name, is_tt, final_match, yn, matching_text, llm_reason, accepted])
        return rows

    def export_ticked_only(self, model: str):
        stamp = time.strftime("%Y%m%d-%H%M%S")
        out = model_exports_dir(model) / f"ticksheet_filtered_{stamp}.csv"
        filt = []
        for row in self.rows:
            # Handle new 9-column format: [nr, name, lv_match, en_match, final_match, include_sym, lv_text, en_text, consensus_reasoning]
            row = (row + [""]*9)[:9]

            nr, name, lv_match, en_match, final_match, include_sym, lv_text, en_text, consensus_reasoning = row[:9]

            # Only export ticked items that are not titles
            if include_sym == CHECKED and "—" not in str(name):
                is_tt = "N"  # Not a title
                matching_text = lv_text or en_text or ""
                llm_reason = consensus_reasoning or ""
                # Export format: [Nr Code, Variable Name, Is TT, Match, Matching Text, LLM_Reason, Accepted]
                filt.append([nr, name, is_tt, final_match, matching_text, llm_reason, "X"])

        # Use proper headers for export
        headers = ["Nr Code", "Variable Name", "Is TT", "Match (Yes/No/Maybe)", "Matching Text", "LLM_Reason", "Accepted"]
        write_csv(out, headers, filt)
        return out

    def export_reviewed_with_tick(self, model: str):
        stamp = time.strftime("%Y%m%d-%H%M%S")
        out = model_exports_dir(model) / f"ticksheet_reviewed_{stamp}.csv"
        rows = self._rows_for_export_with_accepted()
        write_csv(out,
                  ["Nr Code","Variable Name","Is TT","Match (Yes/No/Maybe)","Include","Matching Text","LLM_Reason","Accepted"],
                  rows)
        return out

    def export_comprehensive_xlsx(self, model: str):
        """Export comprehensive Excel file with all data and explanations"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showwarning("OpenPyXL not installed",
                                   "To export .xlsx, install openpyxl:\n\npip install openpyxl")
            return None

        stamp = time.strftime("%Y%m%d-%H%M%S")
        out = model_exports_dir(model) / f"ticksheet_comprehensive_{stamp}.xlsx"

        wb = Workbook()
        ws = wb.active
        
        # Validate worksheet creation
        if ws is None:
            messagebox.showerror("Export Error", "Failed to create Excel worksheet. Please try again.")
            return None
            
        ws.title = "Comprehensive Analysis"

        # Enhanced headers with detailed information for 9-column structure
        headers = [
            "Nr", "Variable Name", "LV Match", "EN Match", "Final Match",
            "Include", "LV Text", "EN Text",
            "Consensus Reasoning", "Status", "Export Mark"
        ]

        # Style the headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        try:
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                if cell is not None:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to style headers: {str(e)}")
            return None

        # Add data rows
        row_num = 2
        for row in self.rows:
            # Use current 9-column structure: [nr, name, lv_match, en_match, final_match, include_sym, lv_text, en_text, consensus_reasoning]
            row = (row + [""]*9)[:9]
            nr, name, lv_match, en_match, final_match, include_sym, lv_text, en_text, consensus_reasoning = row

            # Determine status and export mark
            if "—" in str(name):
                status = "SECTION HEADER"
                export_mark = ""
            else:
                status = "INCLUDED" if include_sym == CHECKED else "EXCLUDED"
                export_mark = "X" if include_sym == CHECKED else ""

            # Write row data - removed confidence column to match 9-column structure
            data_row = [nr, name, lv_match, en_match, final_match,
                       "✓" if include_sym == CHECKED else "☐",
                       lv_text, en_text, consensus_reasoning, status, export_mark]

            for col, value in enumerate(data_row, 1):
                try:
                    cell = ws.cell(row=row_num, column=col, value=str(value) if value else "")
                    if cell is not None:
                        # Style section headers
                        if "—" in str(name):
                            cell.font = Font(bold=True, size=12)
                            cell.fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
                        # Style included items
                        elif include_sym == CHECKED:
                            cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                except Exception as e:
                    messagebox.showerror("Export Error", f"Failed to write data row {row_num}: {str(e)}")
                    return None

            row_num += 1

        # Auto-adjust column widths
        try:
            for col in range(1, len(headers) + 1):
                max_length = len(headers[col-1])
                for row in range(2, row_num):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))

                # Set column width with reasonable limits
                width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[get_column_letter(col)].width = width
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to adjust column widths: {str(e)}")
            return None

        # Add summary sheet
        try:
            summary_ws = wb.create_sheet("Summary")
            if summary_ws is not None:
                summary_cell = summary_ws.cell(1, 1, "Export Summary")
                if summary_cell is not None:
                    summary_cell.font = Font(bold=True, size=14)
                summary_ws.cell(2, 1, f"Model: {model}")
                summary_ws.cell(3, 1, f"Export Date: {time.strftime('%Y-%m-%d %H:%M:%S')}")
                summary_ws.cell(4, 1, f"Total Features: {len([r for r in self.rows if '—' not in str(r[1])])}")
                summary_ws.cell(5, 1, f"Included Features: {len([r for r in self.rows if r[5] == CHECKED])}")
                summary_ws.cell(6, 1, f"Excluded Features: {len([r for r in self.rows if r[5] == UNCHECKED and '—' not in str(r[1])])}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to create summary sheet: {str(e)}")
            return None

        try:
            wb.save(out)
            return out
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to save Excel file: {str(e)}")
            return None

    def copy_export_column(self):
        """Generate and copy export column (X marks) to clipboard"""
        export_marks = []
        for row in self.rows:
            row = (row + [""]*9)[:9]  # Fixed: Use 9-column structure
            nr, name, lv_match, en_match, final_match, include_sym, lv_text, en_text, consensus_reasoning = row

            if "—" in str(name):
                export_marks.append("")  # Empty for section headers
            else:
                export_marks.append("X" if include_sym == CHECKED else "")

        # Join with newlines for copy-paste
        export_text = "\n".join(export_marks)

        # Copy to clipboard
        try:
            import tkinter as tk
            root = tk.Tk()
            root.withdraw()  # Hide the window
            root.clipboard_clear()
            root.clipboard_append(export_text)
            root.update()  # Now it stays on the clipboard after the window is destroyed
            root.destroy()
            return len([mark for mark in export_marks if mark == "X"])
        except Exception as e:
            print(f"Clipboard error: {e}")
            return 0

    def export_sequential_mega_results(self, model: str):
        """Export Sequential MEGA results with X numbering for ticked items"""
        stamp = time.strftime("%Y%m%d-%H%M%S")
        out = model_exports_dir(model) / f"ticksheet_sequential_mega_export_{stamp}.csv"

        export_rows = []
        for row in self.rows:
            # New 9-column format: [nr, name, lv_match, en_match, final_match, include_sym, lv_text, en_text, consensus_reasoning]
            row = (row + [""]*9)[:9]  # Pad to 9 columns

            nr, name, lv_match, en_match, final_match, include_sym, lv_text, en_text, consensus_reasoning = row

            if "—" in str(name):
                # Section headers
                is_tt = "Y"
                clean_name = name.replace("—", "").strip().upper()
                export_rows.append([nr, clean_name, is_tt, "", "", "", "", "", "", "", ""])
            else:
                # Feature rows
                is_tt = "N"

                # Create accepted column with X for ticked items
                accepted = "X" if include_sym == CHECKED else ""
                include_yn = "Y" if include_sym == CHECKED else "N"

                # Export in Sequential MEGA format: Nr Code, Variable Name, Is TT, LV Match, LV Text, EN Match, EN Text, Consensus Vote, Consensus Text, Consensus Reasoning, Include
                consensus_vote = final_match  # No confidence breakdown anymore

                export_rows.append([
                    nr, name, is_tt, lv_match, lv_text, en_match, en_text,
                    consensus_vote, "", consensus_reasoning, accepted
                ])

        # Sequential MEGA export headers
        headers = [
            "Nr Code", "Variable Name", "Is TT",
            "LV Match", "LV Text", "EN Match", "EN Text",
            "Consensus Vote", "Consensus Text", "Consensus Reasoning", "Include"
        ]

        write_csv(out, headers, export_rows)
        return out

def save_snapshot(model: str, table: ReviewTable, status_var: tk.StringVar):
    wp = working_path(model)

    # Check if file is locked/being used
    try:
        # Try to open file for writing to check if it's locked
        with open(wp, 'w', newline='', encoding='utf-8') as test_file:
            pass  # Just testing if we can write
    except PermissionError:
        status_var.set(f"❌ Cannot save - file is locked: {wp.name}")
        messagebox.showerror("File Locked", f"Cannot save to {wp.name}\n\nThe file may be open in Excel or another program.\nPlease close it and try again.")
        return
    except Exception as e:
        status_var.set(f"❌ Save error: {str(e)}")
        return

    rows = []
    for row in table.current_rows_ui():
        # Handle new 10-column format: [nr, name, lv_match, en_match, final_match, confidence, include_sym, lv_text, en_text, consensus_reasoning]
        row = (row + [""]*10)[:10]
        nr, name, lv_match, en_match, final_match, confidence, include_sym, lv_text, en_text, consensus_reasoning = row

        if "—" in str(name):  # Title rows
            is_tt = "Y"
            yn = ""
            accepted = ""
            match = ""
            snippet = ""
            reason = ""
        else:
            is_tt = ""
            yn = "Y" if include_sym == CHECKED else "N"
            accepted = "X" if include_sym == CHECKED else ""
            match = final_match or en_match or lv_match or ""
            snippet = en_text or lv_text or ""
            reason = consensus_reasoning or ""

        rows.append([nr, name, is_tt, match, yn, snippet, reason, accepted])

    try:
        write_csv(wp,
                  ["Nr Code","Variable Name","Is TT","Match (Yes/No/Maybe)","Include","Matching Text","LLM_Reason","Accepted"],
                  rows)
        status_var.set(f"Saved snapshot → {wp.name}")
    except PermissionError:
        status_var.set(f"❌ Permission denied saving to: {wp.name}")
        messagebox.showerror("Permission Error", f"Cannot write to {wp.name}\n\nPlease ensure:\n1. File is not open in Excel\n2. You have write permissions\n3. Antivirus is not blocking the file")
    except Exception as e:
        status_var.set(f"❌ Save failed: {str(e)}")
        messagebox.showerror("Save Error", f"Failed to save snapshot: {str(e)}")

def merge_rows_keep_yes_maybe(old_ui11, new_rows6):
    merged = []
    for old, new in zip(old_ui11, new_rows6):
        # Handle both old (7-column) and new (11-column) formats
        if len(old) <= 7:
            # Old format: [nr, name, is_tt, match, include_sym, snippet, reason]
            old = (old + [""]*7)[:7]
            nr, name, is_tt = old[:3]
            old_match, old_inc, old_snip, old_reason = old[3], old[4], old[5], old[6]
        else:
            # New 11-column format: [nr, name, is_tt, lv_match, en_match, final_match, include_sym, lv_snippet, en_snippet, lv_reason, en_reason]
            old = (old + [""]*11)[:11]
            nr, name, is_tt = old[:3]
            old_match, old_inc = old[5], old[6]  # final_match, include_sym
            old_snip = old[8] or old[7]  # en_snippet or lv_snippet
            old_reason = old[10] or old[9]  # en_reason or lv_reason

        new = (new + [""]*6)[:6]   # [nr, name, is_tt, match, snippet, reason]
        nr, name, is_tt, new_match, new_snip, new_reason = new

        if (is_tt or "").upper() == "Y":
            # Convert to new 11-column format for title rows
            merged.append([nr, f"— {name.upper()} —", is_tt, "", "", "", "", "", "", "", ""])
            continue

        # prefer your reviewed Yes/Maybe (and your tick + snippet/reason)
        if old_match in ("Yes","Maybe"):
            # Keep old result, convert to new 11-column format
            include_sym = old_inc if old_inc in [CHECKED, UNCHECKED] else (CHECKED if old_match in ("Yes","Maybe") else UNCHECKED)
            merged.append([nr, name, is_tt, "", "", old_match, include_sym, "", old_snip, "", old_reason])
        else:
            # Use new result, convert to new 11-column format
            nm, ns = evidence_sanity_check(new_match, new_snip, name)
            inc_sym = CHECKED if nm in ("Yes","Maybe") else UNCHECKED
            merged.append([nr, name, is_tt, "", "", nm, inc_sym, "", ns, "", new_reason])
    return merged

def run_pipeline(model: str, text: str, status_var: tk.StringVar, run_btn: tk.Button, table: ReviewTable, llm_model: str = "gpt-4o"):
    def worker():
        try:
            status_var.set("Saving input…")
            SESSIONS_DIR.mkdir(parents=True, exist_ok=True)
            specfile = SESSIONS_DIR / f"{model}_input.txt"
            specfile.write_text(text, encoding="utf-8")
            cmd = [sys.executable, str(ROOT / "run_pipeline.py"),
                   "--model", model, "--specfile", str(specfile),
                   "--master", str(MASTERLIST_DIR / f"{model}.csv"),
                   "--exportdir", str(EXPORTS_DIR), "--llm", llm_model]
            status_var.set("Running pipeline…")
            proc = subprocess.run(cmd, capture_output=True, text=True)
            if proc.returncode != 0:
                msg = proc.stderr.strip() or proc.stdout.strip() or "Unknown error."
                status_var.set("❌ Failed")
                messagebox.showerror("Pipeline error", msg)
                return
            status_var.set("Loading results…")
            latest = find_latest_ticksheet(model)
            if not latest:
                status_var.set("❌ No results found")
                messagebox.showerror("Not found", "No ticksheet found in exports.")
                return
            header, data = read_csv_any(latest)
            header_lower = [(h or "").strip().lower() for h in header]
            if "include" in header_lower: table.load_from_reviewed_rows(data)
            else: table.load_from_pipeline_rows(data)
            save_snapshot(model, table, status_var)
            messagebox.showinfo("Loaded", f"Loaded {latest.name}\nAutosaved snapshot as ticksheet_working.csv")
        except Exception as e:
            status_var.set("❌ Failed")
            messagebox.showerror("Error", str(e))
        finally:
            run_btn.config(state=tk.NORMAL)
    run_btn.config(state=tk.DISABLED)
    threading.Thread(target=worker, daemon=True).start()

def rerun_pipeline_merge(model: str, text: str, status_var: tk.StringVar, run_btn: tk.Button, table: ReviewTable, llm_model: str = "gpt-4o"):
    def worker():
        try:
            status_var.set("Saving input…")
            SESSIONS_DIR.mkdir(parents=True, exist_ok=True)
            specfile = SESSIONS_DIR / f"{model}_input.txt"
            specfile.write_text(text, encoding="utf-8")
            cmd = [sys.executable, str(ROOT / "run_pipeline.py"),
                   "--model", model, "--specfile", str(specfile),
                   "--master", str(MASTERLIST_DIR / f"{model}.csv"),
                   "--exportdir", str(EXPORTS_DIR), "--llm", llm_model]
            status_var.set("Rerunning pipeline…")
            proc = subprocess.run(cmd, capture_output=True, text=True)
            if proc.returncode != 0:
                msg = proc.stderr.strip() or proc.stdout.strip() or "Unknown error."
                status_var.set("❌ Failed")
                messagebox.showerror("Pipeline error", msg)
                return
            latest = find_latest_ticksheet(model)
            if not latest:
                status_var.set("❌ No results found")
                messagebox.showerror("Not found", "No ticksheet found in exports.")
                return
            header, new_data = read_csv_any(latest)
            header_lower = [(h or "").strip().lower() for h in header]
            if "include" in header_lower:
                inc_i = header_lower.index("include")
                new_rows6 = [[c for j,c in enumerate(r) if j != inc_i] for r in new_data]
                if "accepted" in header_lower:
                    acc_i = header_lower.index("accepted")
                    new_rows6 = [[c for j,c in enumerate(r) if j != acc_i] for r in new_rows6]
            else:
                new_rows6 = new_data
            old_ui11 = table.current_rows_ui()
            if not old_ui11: table.load_from_pipeline_rows(new_rows6)
            else: table.set_rows_ui(merge_rows_keep_yes_maybe(old_ui11, new_rows6))
            save_snapshot(model, table, status_var)
            messagebox.showinfo("Merged", "Rerun complete.\nKept your Yes/Maybe rows and ticks.\nAutosaved snapshot.")
        except Exception as e:
            status_var.set("❌ Failed")
            messagebox.showerror("Error", str(e))
        finally:
            run_btn.config(state=tk.NORMAL)
    run_btn.config(state=tk.DISABLED)
    threading.Thread(target=worker, daemon=True).start()

def run_dual_pipeline(model: str, text: str, status_var: tk.StringVar, run_btn: tk.Button, table: ReviewTable, llm_model: str = "gpt-4o"):
    """Run dual-language pipeline (Latvian + English) with merged results"""
    def worker():
        try:
            status_var.set("Saving input…")
            SESSIONS_DIR.mkdir(parents=True, exist_ok=True)
            specfile = SESSIONS_DIR / f"{model}_input.txt"
            specfile.write_text(text, encoding="utf-8")

            # Check if English masterlist exists
            master_lv = MASTERLIST_DIR / f"{model}.csv"
            master_en = MASTERLIST_DIR / f"{model}_en.csv"

            if not master_lv.exists():
                status_var.set("❌ Failed")
                messagebox.showerror("Missing File", f"Latvian masterlist not found: {master_lv.name}")
                return

            if not master_en.exists():
                status_var.set("❌ Failed")
                messagebox.showerror("Missing File",
                    f"English masterlist not found: {master_en.name}\n\n"
                    f"Please create {master_en.name} with English feature names in the masterlists folder.")
                return

            # Check if English prompt exists
            prompt_en = ROOT / "prompt_en.txt"
            if not prompt_en.exists():
                status_var.set("❌ Failed")
                messagebox.showerror("Missing File",
                    "English prompt file not found: prompt_en.txt\n\n"
                    "This should have been created automatically.")
                return

            cmd = [sys.executable, str(ROOT / "run_dual_pipeline.py"),
                   "--model", model,
                   "--specfile", str(specfile),
                   "--master-lv", str(master_lv),
                   "--master-en", str(master_en),
                   "--prompt-lv", str(ROOT / "prompt.txt"),
                   "--prompt-en", str(prompt_en),
                   "--exportdir", str(EXPORTS_DIR)]

            status_var.set("Running dual-language pipeline…")
            proc = subprocess.run(cmd, capture_output=True, text=True)

            if proc.returncode != 0:
                msg = proc.stderr.strip() or proc.stdout.strip() or "Unknown error."
                status_var.set("❌ Failed")
                messagebox.showerror("Pipeline error", msg)
                return

            status_var.set("Loading dual-language results…")

            # Find the latest result files
            model_dir = EXPORTS_DIR / model

            # Find dual, LV, and EN files
            dual_files = list(model_dir.glob("ticksheet_dual_*.csv"))
            lv_files = list(model_dir.glob("ticksheet_LV_*.csv"))
            en_files = list(model_dir.glob("ticksheet_EN_*.csv"))

            if not dual_files:
                status_var.set("❌ No dual results found")
                messagebox.showerror("Not found", "Could not find dual-language result file (ticksheet_dual_*.csv).")
                return

            # Get the most recent dual file
            latest_dual = max(dual_files, key=lambda p: p.stat().st_mtime)

            # Load dual-language results directly
            _, dual_data = read_csv_any(latest_dual)

            # Load into dual-language view using the new format
            table.load_from_dual_csv(dual_data)
            save_snapshot(model, table, status_var)

            messagebox.showinfo("Dual-Language Complete",
                f"Loaded dual results: {latest_dual.name}\n"
                f"Sequential Latvian + English recognition.\n"
                f"Autosaved snapshot as ticksheet_working.csv")

        except Exception as e:
            status_var.set("❌ Failed")
            messagebox.showerror("Error", str(e))
        finally:
            run_btn.config(state=tk.NORMAL)

    run_btn.config(state=tk.DISABLED)
    threading.Thread(target=worker, daemon=True).start()

def main():
    ROOT.mkdir(exist_ok=True)
    MASTERLIST_DIR.mkdir(exist_ok=True)
    SESSIONS_DIR.mkdir(exist_ok=True)
    EXPORTS_DIR.mkdir(exist_ok=True)
    models = list_models()

    # Modern color scheme inspired by contemporary web design (Inchcape style)
    COLORS = {
        'bg_primary': '#FFFFFF',       # Pure white background
        'bg_secondary': '#F8FAFC',     # Very light slate gray
        'bg_accent': '#F1F5F9',        # Light slate for elevated surfaces
        'text_primary': '#1E293B',     # Dark slate text
        'text_secondary': '#64748B',   # Medium slate text
        'text_muted': '#94A3B8',       # Muted slate text
        'blue_primary': '#1E3A8A',     # Deep professional blue
        'blue_secondary': '#3B82F6',   # Bright blue for accents
        'green_primary': '#059669',    # Modern emerald green
        'red_primary': '#DC2626',      # Clean red
        'orange_primary': '#EA580C',   # Vibrant orange
        'border_light': '#E2E8F0',     # Light border
        'border_strong': '#CBD5E1',    # Stronger border
        'shadow': '#00000015',         # Subtle shadow
        'row_even': '#FFFFFF',         # Pure white
        'row_odd': '#F8FAFC',          # Very light gray
        'hover': '#F1F5F9',            # Hover state
        'selected': '#EFF6FF',         # Selection background
    }

    # Modern typography using system fonts with improved readability
    FONTS = {
        'heading': ('Segoe UI', 14, 'bold'),      # Increased from 13 to 14
        'subheading': ('Segoe UI', 12, 'bold'),   # Increased from 11 to 12
        'body': ('Segoe UI', 11),                 # Increased from 10 to 11
        'small': ('Segoe UI', 10),                # Increased from 9 to 10
        'code': ('Segoe UI', 10),                 # Increased from 9 to 10
        'button': ('Segoe UI', 11, 'bold'),       # Increased from 10 to 11
        'button_small': ('Segoe UI', 10, 'bold'), # Increased from 9 to 10
    }

    # Modern button styling function
    def style_modern_button(button, button_type='primary'):
        """Apply modern styling to buttons for contemporary appearance"""
        if button_type == 'primary':
            button.configure(
                bg=COLORS['blue_primary'],
                fg='#FFFFFF',
                font=FONTS['button'],
                relief='flat',
                borderwidth=0,
                padx=16,
                pady=8,
                cursor='hand2',
                activebackground='#1E40AF',  # Darker blue on hover
                activeforeground='#FFFFFF'
            )
        elif button_type == 'secondary':
            button.configure(
                bg=COLORS['bg_accent'],
                fg=COLORS['text_primary'],
                font=FONTS['button'],
                relief='flat',
                borderwidth=1,
                padx=16,
                pady=8,
                cursor='hand2',
                activebackground='#E2E8F0',
                activeforeground=COLORS['text_primary']
            )
        elif button_type == 'small':
            button.configure(
                bg=COLORS['blue_primary'],
                fg='#FFFFFF',
                font=FONTS['button_small'],
                relief='flat',
                borderwidth=0,
                padx=12,
                pady=6,
                cursor='hand2',
                activebackground='#1E40AF',
                activeforeground='#FFFFFF'
            )

    # Modern window setup with contemporary styling
    win = tk.Tk()
    win.title("🚗 Spec Matcher Pro — AI-Powered Dual Language Review")

    # Set to maximized with modern styling
    win.state('zoomed')
    win.configure(bg=COLORS['bg_primary'])

    # Modern window styling
    try:
        # Try to set modern window appearance on Windows
        win.tk.call('tk', 'scaling', 1.0)  # Proper DPI scaling
    except:
        pass

    # Allow user to toggle fullscreen
    win.bind('<F11>', lambda e: win.attributes('-fullscreen', not win.attributes('-fullscreen')))
    win.bind('<Escape>', lambda e: win.attributes('-fullscreen', False))

    # Set minimum size for when not maximized
    win.minsize(1000, 700)

    # Apply better styling to window
    win.configure(bg=COLORS['bg_primary'])

    # Set window icon (if available)
    try:
        win.iconbitmap(default="favicon.ico")  # Optional: add if you have an icon
    except:
        pass

    # Top menu (export fallback so it never 'disappears')
    menubar = tk.Menu(win)
    filemenu = tk.Menu(menubar, tearoff=0)
    def _export_comprehensive_menu():
        if not table.rows: return messagebox.showwarning("Nothing to export", "Run analysis first.")
        out = table.export_comprehensive_xlsx(model_var.get().strip())
        if out: messagebox.showinfo("Exported", f"Comprehensive Excel saved: {out.name}")
    def _export_copy_menu():
        if not table.rows: return messagebox.showwarning("Nothing to copy", "Run analysis first.")
        count = table.copy_export_column()
        messagebox.showinfo("Copied", f"Export column copied ({count} X marks)")
    def _export_csv_menu():
        if not table.rows: return messagebox.showwarning("Nothing to export", "Run analysis first.")
        out = table.export_reviewed_with_tick(model_var.get().strip())
        messagebox.showinfo("Exported", f"CSV saved: {out.name}")

    filemenu.add_command(label="📊 Export Comprehensive Excel", command=_export_comprehensive_menu)
    filemenu.add_command(label="📋 Copy Export Column", command=_export_copy_menu)
    filemenu.add_separator()
    filemenu.add_command(label="📄 Export CSV (Legacy)", command=_export_csv_menu)
    menubar.add_cascade(label="File", menu=filemenu)
    win.config(menu=menubar)

    # Top bar
    # Modern top bar with clean design and professional styling
    top = tk.Frame(win, bg=COLORS['bg_accent'], relief="flat", bd=0)
    top.pack(fill="x", padx=10, pady=8)

    # Model selection with modern styling
    model_frame = tk.Frame(top, bg=COLORS['bg_accent'])
    model_frame.pack(side="left")
    tk.Label(model_frame, text="🚗 Car Model:", font=FONTS['body'], bg=COLORS['bg_accent'],
             fg=COLORS['text_primary']).pack(side="left")
    model_var = tk.StringVar(value=models[0] if models else "")
    model_menu = tk.OptionMenu(model_frame, model_var, *(models or [""]))
    model_menu.config(font=FONTS['body'], bg=COLORS['bg_primary'], fg=COLORS['text_primary'],
                      relief="flat", bd=1, activebackground=COLORS['blue_primary'])
    model_menu.pack(side="left", padx=(8, 12))

    # ChatGPT model selection with modern design
    llm_frame = tk.Frame(top, bg=COLORS['bg_accent'])
    llm_frame.pack(side="left")
    tk.Label(llm_frame, text="🤖 ChatGPT Model:", font=FONTS['body'], bg=COLORS['bg_accent'],
             fg=COLORS['text_primary']).pack(side="left")
    llm_var = tk.StringVar(value="gpt-4o")
    llm_options = ["gpt-4o", "gpt-4o-mini"]
    llm_menu = tk.OptionMenu(llm_frame, llm_var, *llm_options)
    llm_menu.config(font=FONTS['body'], bg=COLORS['bg_primary'], fg=COLORS['text_primary'],
                    relief="flat", bd=1, width=12, activebackground=COLORS['blue_primary'])
    llm_menu.pack(side="left", padx=(8, 12))

    # Text normalization option with modern styling
    normalize_var = tk.BooleanVar(value=True)  # Default to enabled
    normalize_cb = tk.Checkbutton(top, text="🧹 Normalize Text", variable=normalize_var,
                                 font=FONTS['body'], bg=COLORS['bg_accent'], fg=COLORS['text_primary'],
                                 activebackground=COLORS['bg_accent'], selectcolor=COLORS['bg_primary'])
    normalize_cb.pack(side="left", padx=(8, 12))

    # Enhanced status with modern color coding
    status_frame = tk.Frame(top, bg=COLORS['bg_accent'])
    status_frame.pack(side="right")
    status_var = tk.StringVar(value="Ready")
    status_label = tk.Label(status_frame, textvariable=status_var, font=FONTS['body'],
                           bg=COLORS['bg_primary'], fg=COLORS['text_primary'], padx=12, pady=4,
                           relief="flat", bd=1)
    status_label.pack(side="right")

    # Function to update status with colors
    def update_status(message, color="#323130", bg_color="#E1DFDD"):
        original_set(message)  # Use original set method to avoid recursion
        status_label.config(fg=color, bg=bg_color)

    # Override status_var.set to use color coding
    original_set = status_var.set
    def enhanced_set(value):
        if "❌" in value or "Failed" in value or "Error" in value:
            update_status(value, "#D13438", "#FDE7E9")  # Red
        elif "✓" in value or "Complete" in value or "Success" in value:
            update_status(value, "#107C10", "#DFF6DD")  # Green
        elif "Running" in value or "Loading" in value:
            update_status(value, "#8764B8", "#F3F0FF")  # Purple
        else:
            update_status(value, "#323130", "#E1DFDD")  # Default
    status_var.set = enhanced_set

    # Modern paste area with clean contemporary design
    mid = tk.Frame(win, bg=COLORS['bg_primary'], relief="flat", bd=0)
    mid.pack(fill="x", padx=10, pady=(0, 8))

    text_label = tk.Label(mid, text="📝 Paste spec/features text:", font=FONTS['subheading'],
                         bg=COLORS['bg_primary'], fg=COLORS['text_primary'])
    text_label.pack(anchor="w")

    # Text area with modern styling and clean borders
    text_frame = tk.Frame(mid, bg=COLORS['bg_primary'])
    text_frame.pack(fill="x", pady=(4, 12))

    text_box = tk.Text(text_frame, height=8, wrap="word", font=FONTS['code'],
                      bg=COLORS['bg_primary'], fg=COLORS['text_primary'], relief="flat", bd=1,
                      selectbackground=COLORS['blue_primary'], selectforeground="#FFFFFF",
                      insertbackground=COLORS['text_primary'])
    text_box.pack(fill="x", padx=1, pady=1)

    # Add modern border effect with subtle gray
    text_frame.config(bg=COLORS['text_secondary'], bd=1, relief="solid")

    btn_row = tk.Frame(mid, bg=COLORS['bg_primary'], relief="flat", bd=0)
    btn_row.pack(fill="x")
    def on_load_file():
        path = filedialog.askopenfilename(title="Open text file", filetypes=[("Text files","*.txt"), ("All files","*.*")])
        if not path: return
        text_box.delete("1.0", tk.END)
        text_box.insert("1.0", pathlib.Path(path).read_text(encoding="utf-8", errors="ignore"))

    def on_load_dual_csv():
        path = filedialog.askopenfilename(title="Open dual-language CSV",
                                         filetypes=[("CSV files","*.csv"), ("All files","*.*")],
                                         initialdir=str(EXPORTS_DIR))
        if not path: return
        try:
            _, dual_data = read_csv_any(pathlib.Path(path))
            table.load_from_dual_csv(dual_data)
            status_var.set("Loaded dual CSV")
            messagebox.showinfo("Loaded", f"Loaded dual-language results from {pathlib.Path(path).name}")
        except Exception as e:
            messagebox.showerror("Load Error", f"Failed to load CSV: {str(e)}")

    def auto_commit_to_dictionary(results_data, analysis_type="analysis"):
        """Automatically commit high-confidence results to dictionary for learning"""
        try:
            from feature_dictionary import FeatureDictionary
            fd = FeatureDictionary()

            committed_count = 0
            high_confidence_threshold = 0.8  # Only auto-commit very confident matches

            for item in results_data:
                try:
                    # Only auto-commit high-confidence "Yes" matches
                    if (item.get('Match Status') == 'Yes' and
                        isinstance(item.get('Confidence'), (int, float)) and
                        item.get('Confidence') >= high_confidence_threshold):

                        # Add to dictionary as a verified good match
                        feature_id = fd.add_feature(
                            original_text=item.get('Original Text', ''),
                            language='EN',  # Default language
                            model_name=llm_var.get(),
                            matched_code=item.get('Code', ''),
                            matched_name=item.get('Feature Name', ''),
                            match_status='Yes',
                            confidence=item.get('Confidence', 0.8),
                            reasoning=f"Auto-committed from {analysis_type} - high confidence match",
                            method="auto_commit",
                            check_similarity=True  # Avoid duplicates
                        )

                        if feature_id:
                            # Mark as verified since it's high confidence
                            fd.mark_as_verified(feature_id, True,
                                              f"Auto-verified from {analysis_type} - confidence >= 80%")
                            committed_count += 1

                except Exception as e:
                    print(f"Error auto-committing item: {str(e)}")
                    continue

            if committed_count > 0:
                print(f"Auto-committed {committed_count} high-confidence matches to dictionary")

        except Exception as e:
            print(f"Error in auto-commit: {str(e)}")

    # LEGACY CODE: This function is no longer used in the UI (button commented out)
    # Can be safely removed after confirming no external dependencies
    # Replaced by: on_run_learning_mega() which provides enhanced functionality
    def on_run_intelligent():
        model = model_var.get().strip()
        if not model: return messagebox.showwarning("Missing model", "No masterlist found. Add one to masterlists/.")
        text = text_box.get("1.0", tk.END).strip()
        if not text: return messagebox.showwarning("Empty text", "Paste spec text or use 'Load .txt'.")
        llm_model = llm_var.get().strip()

        def worker():
            try:
                status_var.set("Running intelligent matcher...")
                # intelligent_btn.config(state=tk.DISABLED)  # Button commented out

                # Save input text
                SESSIONS_DIR.mkdir(parents=True, exist_ok=True)
                specfile = SESSIONS_DIR / f"{model}_input.txt"
                specfile.write_text(text, encoding="utf-8")

                # Import and use intelligent_matcher directly
                from intelligent_matcher import intelligent_match
                from pathlib import Path

                # Setup paths
                master_path = MASTERLIST_DIR / f"{model}.csv"
                export_dir = EXPORTS_DIR / model
                export_dir.mkdir(parents=True, exist_ok=True)

                timestamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
                output_path = export_dir / f"ticksheet_intelligent_{timestamp}.csv"

                # Run intelligent matching with normalization setting
                use_normalization = normalize_var.get()
                results = intelligent_match(text, master_path, output_path, llm_model, use_normalization)

                if results:
                    # Load the results
                    _, data = read_csv_any(output_path)
                    table.load_from_pipeline_rows(data)
                    save_snapshot(model, table, status_var)

                    # Auto-commit disabled - use "Learn from results" button instead
                    # try:
                    #     auto_commit_to_dictionary(data, "intelligent_analysis")
                    # except Exception as e:
                    #     print(f"Auto-commit failed: {str(e)}")

                    # Success message with normalization status
                    norm_status = "with text normalization" if use_normalization else "without text normalization"
                    messagebox.showinfo("Intelligent Match Complete",
                        f"✅ Intelligent matching completed {norm_status}!\n\n"
                        f"📊 Processed {len(results)} masterlist items\n"
                        f"🤖 Used {llm_model} for optimal results\n"
                        f"💾 Results saved to: {output_path.name}\n"
                        f"📋 Auto-saved snapshot as ticksheet_working.csv\n"
                        f"🧠 High-confidence matches added to learning dictionary")
                else:
                    status_var.set("❌ No results generated")
                    messagebox.showerror("Error", "No results were generated from intelligent matching")

            except Exception as e:
                status_var.set("❌ Failed")
                messagebox.showerror("Error", f"Intelligent matching failed: {str(e)}")
            finally:
                # intelligent_btn.config(state=tk.NORMAL)  # Button commented out
                pass  # Empty finally block

        threading.Thread(target=worker, daemon=True).start()

    # LEGACY CODE: This function is no longer used in the UI (button commented out)
    # Can be safely removed after confirming no external dependencies
    # Replaced by: on_run_learning_mega() which provides enhanced learning functionality
    def on_run_mega():
        model = model_var.get().strip()
        if not model: return messagebox.showwarning("Missing model", "No masterlist found. Add one to masterlists/.")
        text = text_box.get("1.0", tk.END).strip()
        if not text: return messagebox.showwarning("Empty text", "Paste spec text or use 'Load .txt'.")
        llm_model = llm_var.get().strip()

        # Check for both LV and EN masterlists
        master_lv_path = MASTERLIST_DIR / f"{model}.csv"
        master_en_path = MASTERLIST_DIR / f"{model}_en.csv"

        if not master_lv_path.exists():
            return messagebox.showwarning("Missing LV masterlist", f"Could not find {model}.csv")
        if not master_en_path.exists():
            return messagebox.showwarning("Missing EN masterlist", f"Could not find {model}_en.csv")

        def worker():
            try:
                status_var.set("Running Sequential MEGA...")
                # mega_btn.config(state=tk.DISABLED)  # Button commented out

                # Progress callback for detailed status updates
                def progress_update(description):
                    status_var.set(description)
                    win.update_idletasks()

                # Save input text
                SESSIONS_DIR.mkdir(parents=True, exist_ok=True)
                specfile = SESSIONS_DIR / f"{model}_input.txt"
                specfile.write_text(text, encoding="utf-8")

                # Import and use sequential mega matcher
                from sequential_mega_matcher import sequential_mega_match
                from pathlib import Path

                # Setup paths
                export_dir = EXPORTS_DIR / model
                export_dir.mkdir(parents=True, exist_ok=True)

                timestamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
                output_path = export_dir / f"ticksheet_sequential_mega_{timestamp}.csv"

                # Run Sequential MEGA matching with progress callback
                use_normalization = normalize_var.get()
                results = sequential_mega_match(text, model, output_path, llm_model, use_normalization, progress_update)

                if results:
                    # Load the results (sequential MEGA CSV has 11 columns now)
                    _, data = read_csv_any(output_path)
                    # Convert sequential format to UI format
                    ui_data = []
                    for row in data:
                        if len(row) >= 11:
                            # Sequential format: Nr Code, Variable Name, Is TT, LV Match, LV Text, EN Match, EN Text, Consensus Vote, Consensus Text, Consensus Reasoning, Include
                            nr_code, name, is_tt, lv_match, lv_text, en_match, en_text, consensus_vote, consensus_text, consensus_reasoning, include = row[:11]

                            # Map to dual-language UI format (10 columns)
                            ui_row = [
                                nr_code, name, is_tt,           # 0,1,2: Code, Name, TT
                                lv_match, en_match, consensus_vote.split(' ')[0],  # 3,4,5: LV Match, EN Match, Final Match
                                include,                        # 6: Include
                                lv_text, en_text,              # 7,8: LV Text, EN Text
                                consensus_reasoning             # 9: Consensus Reasoning
                            ]
                            ui_data.append(ui_row)

                    table.load_from_dual_csv(ui_data)
                    save_snapshot(model, table, status_var)

                    # Success message with normalization status
                    norm_status = "with text normalization" if use_normalization else "without text normalization"
                    messagebox.showinfo("Sequential MEGA Complete",
                        f"🚀 Sequential MEGA matching completed {norm_status}!\n\n"
                        f"🇱� Latvian analysis complete\n"
                        f"🇬🇧 English analysis complete\n"
                        f"🔥 3x Hot consensus complete\n"
                        f"📊 Processed {len([r for r in results if r[2] != 'Y'])} features\n"
                        f"🤖 Used {llm_model} at temp 0.2/0.5\n"
                        f"💾 Results saved to: {output_path.name}\n"
                        f"📋 Auto-saved snapshot as ticksheet_working.csv")
                else:
                    status_var.set("❌ No results generated")
                    messagebox.showerror("Error", "No results were generated from Sequential MEGA matching")

            except Exception as e:
                status_var.set("❌ Failed")
                messagebox.showerror("Error", f"Sequential MEGA matching failed: {str(e)}")
            finally:
                # mega_btn.config(state=tk.NORMAL)  # Button commented out
                pass  # Empty finally block

        threading.Thread(target=worker, daemon=True).start()

    def on_run_learning_mega():
        """Run Learning Sequential MEGA with enhanced dictionary integration"""
        model = model_var.get()
        text = text_box.get("1.0", "end").strip()
        llm_model = llm_var.get().strip()

        if not model:
            messagebox.showwarning("Warning", "Please select a model first")
            return
        if not text:
            messagebox.showwarning("Warning", "Please load some text to process")
            return

        learning_mega_btn.config(state=tk.DISABLED)
        status_var.set("🎓 Starting Learning Sequential MEGA...")

        def worker():
            try:
                from learning_sequential_mega import learning_sequential_mega_match

                # Get normalization preference
                use_normalization = normalize_var.get()
                status_var.set(f"🎓 Learning MEGA: Analyzing with enhanced dictionary knowledge...")

                # Create output path for this run
                timestamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
                output_path = model_exports_dir(model) / f"learning_mega_{timestamp}.csv"

                # Run the learning sequential MEGA pipeline
                results = learning_sequential_mega_match(
                    spec_text=text,
                    base_model=model,
                    output_path=output_path,
                    llm_model=llm_model,
                    normalize_text=use_normalization,
                    progress_callback=lambda msg: status_var.set(f"🎓 Learning MEGA: {msg}")
                )

                if results:
                    # Convert to UI format (11 columns: Nr Code, Variable Name, Is TT, LV Match, EN Match, Final Match, Include, LV Text, EN Text, Consensus Reasoning, Accepted)
                    ui_data = []
                    for i, result in enumerate(results):
                        if len(result) >= 9:  # Ensure we have all expected columns
                            # Convert to include checkbox format
                            final_match = result[5] if len(result) > 5 else "No"
                            include_status = CHECKED if final_match in ("Yes", "Maybe") else UNCHECKED

                            ui_row = [
                                result[0],  # Nr Code
                                result[1],  # Variable Name
                                result[2],  # Is TT
                                result[3],  # LV Match
                                result[4],  # EN Match
                                result[5],  # Final Match
                                include_status,  # Include checkbox
                                result[6] if len(result) > 6 else "",  # LV Text
                                result[7] if len(result) > 7 else "",  # EN Text
                                result[8] if len(result) > 8 else "",  # Consensus Reasoning
                                ""  # Accepted (empty for now)
                            ]
                            ui_data.append(ui_row)

                    # Save results to timestamped CSV with correct headers
                    learning_mega_headers = [
                        "Nr Code", "Variable Name", "Is TT",
                        "LV Match", "EN Match", "Final Match", "Include",
                        "LV Text", "EN Text", "Consensus Reasoning", "Accepted"
                    ]
                    write_csv(output_path, learning_mega_headers, ui_data)

                    # Load into table
                    table.load_from_dual_csv(ui_data)
                    save_snapshot(model, table, status_var)

                    # Success message with learning enhancement info
                    norm_status = "with text normalization" if use_normalization else "without text normalization"
                    messagebox.showinfo("Learning Sequential MEGA Complete",
                        f"🎓 Learning Sequential MEGA matching completed {norm_status}!\n\n"
                        f"🧠 Enhanced with learned dictionary knowledge\n"
                        f"🇱🇻 Latvian analysis with learned patterns\n"
                        f"🇬🇧 English analysis with learned features\n"
                        f"🔥 3x Hot consensus with learning hints\n"
                        f"📊 Processed {len([r for r in results if r[2] != 'Y'])} features\n"
                        f"🤖 Used {llm_model} with enhanced prompts\n"
                        f"💾 Results saved to: {output_path.name}\n"
                        f"📋 Auto-saved snapshot as ticksheet_working.csv\n\n"
                        f"💡 Tip: Use 'Learn from Results' button to improve future accuracy!")
                else:
                    status_var.set("❌ No results generated")
                    messagebox.showerror("Error", "No results were generated from Learning Sequential MEGA matching")

            except Exception as e:
                status_var.set("❌ Failed")
                messagebox.showerror("Error", f"Learning Sequential MEGA matching failed: {str(e)}")
            finally:
                learning_mega_btn.config(state=tk.NORMAL)

        threading.Thread(target=worker, daemon=True).start()

    def show_progress_window(title="Processing...", max_value=100):
        """Show a progress window with progress bar"""
        progress_window = tk.Toplevel(win)
        progress_window.title(title)
        progress_window.geometry("400x150")
        progress_window.resizable(False, False)
        progress_window.transient(win)
        progress_window.grab_set()

        # Center the window
        progress_window.update_idletasks()
        x = (progress_window.winfo_screenwidth() // 2) - (400 // 2)
        y = (progress_window.winfo_screenheight() // 2) - (150 // 2)
        progress_window.geometry(f"400x150+{x}+{y}")

        # Icon and title
        title_frame = tk.Frame(progress_window)
        title_frame.pack(pady=10)

        title_label = tk.Label(title_frame, text="🚀 " + title, font=("Segoe UI", 12, "bold"))
        title_label.pack()

        # Progress label
        progress_label = tk.Label(progress_window, text="Initializing...", font=("Segoe UI", 10))
        progress_label.pack(pady=5)

        # Progress bar
        progress_bar = ttk.Progressbar(progress_window, length=350, mode='determinate', maximum=max_value)
        progress_bar.pack(pady=10)

        # Status label
        status_label = tk.Label(progress_window, text="Please wait...", font=("Segoe UI", 8), fg="#666666")
        status_label.pack()

        # Store references using a dictionary to avoid attribute assignment issues
        progress_data = {
            'window': progress_window,
            'progress_bar': progress_bar,
            'progress_label': progress_label,
            'status_label': status_label
        }

        progress_window.update()
        return progress_data

    def update_progress(progress_data, value, label_text="", status_text=""):
        """Update progress bar and labels"""
        if progress_data and progress_data['window'].winfo_exists():
            progress_data['progress_bar']['value'] = value
            if label_text:
                progress_data['progress_label'].config(text=label_text)
            if status_text:
                progress_data['status_label'].config(text=status_text)
            progress_data['window'].update()

    def close_progress(progress_data):
        """Close progress window safely"""
        if progress_data and progress_data['window'].winfo_exists():
            progress_data['window'].destroy()

    # LEGACY CODE: This function is no longer used in the UI (button commented out)
    # Can be safely removed after confirming no external dependencies
    # Replaced by: on_run_learning_mega() which includes enhanced consensus
    def on_run_consensus():
        """Run consensus matching (3 attempts, majority vote) and merge with existing results"""
        model = model_var.get().strip()
        if not model:
            messagebox.showerror("Error", "Please select a model first.")
            return

        llm_model = llm_var.get().strip()
        # consensus_btn.config(state=tk.DISABLED)  # Button commented out

        def worker():
            progress_data = None
            try:
                # Show progress window
                progress_data = show_progress_window("Smart Consensus Matching", 100)

                update_progress(progress_data, 10, "💾 Saving current results", "Preserving your current matches...")

                # Save current state before running new matching
                current_results = {}
                for row_data in table.rows:
                    if len(row_data) >= 7:  # Make sure we have enough columns
                        nr_code = row_data[0]  # Nr Code
                        final_match = row_data[5]  # Final Match column
                        include = row_data[6]  # Include column

                        # Only save if it's marked as included and has a match
                        if include == CHECKED and final_match and final_match.strip():
                            current_results[nr_code] = final_match.strip()

                update_progress(progress_data, 25, "🚀 Starting consensus matcher", "Running 3 intelligent matches...")

                # Run consensus matcher
                cmd = [sys.executable, "consensus_matcher.py", model, "3", "--llm", llm_model]
                print(f"Running command: {' '.join(cmd)}")  # Debug output
                proc = subprocess.run(cmd, capture_output=True, text=True, encoding='utf-8',
                                    cwd=pathlib.Path.cwd(), errors='replace')

                update_progress(progress_data, 70, "🔄 Processing results", "Analyzing consensus votes...")

                print(f"Return code: {proc.returncode}")  # Debug output
                if proc.stdout:
                    print(f"Stdout: {proc.stdout}")  # Debug output
                if proc.stderr:
                    print(f"Stderr: {proc.stderr}")  # Debug output

                if proc.returncode == 0:
                    update_progress(progress_data, 85, "📊 Loading results", "Merging with existing matches...")

                    # Load consensus results
                    consensus_file = pathlib.Path(f"exports/{model}/consensus_intelligent_match.csv")
                    print(f"Looking for consensus file: {consensus_file}")  # Debug output
                    print(f"File exists: {consensus_file.exists()}")  # Debug output

                    if consensus_file.exists():
                        try:
                            header, data = read_csv_any(consensus_file)
                            print(f"Loaded {len(data)} rows from consensus file")  # Debug output

                            update_progress(progress_data, 95, "🔗 Merging results", "Combining old and new matches...")

                            # Merge with previous results
                            for row in data:
                                if len(row) >= 4:  # Make sure we have enough columns
                                    nr_code = row[0]  # Nr Code
                                    new_match_status = row[3]  # Match (Yes/No/Maybe)
                                    new_match_text = row[4] if len(row) > 4 else ""  # Matching Text

                                    # If this row has a previous result and new result found
                                    if (nr_code in current_results and
                                        new_match_status in ['Yes', 'Maybe'] and
                                        new_match_text and new_match_text.strip()):

                                        # Combine with "/" separator
                                        combined_text = f"{current_results[nr_code]} / {new_match_text.strip()}"
                                        row[4] = combined_text  # Update the matching text
                                        print(f"Merged {nr_code}: {combined_text}")

                            table.load_from_pipeline_rows(data)
                            save_snapshot(model, table, status_var)

                            # Auto-commit disabled - use "Learn from results" button instead
                            # try:
                            #     auto_commit_to_dictionary(data, "consensus_analysis")
                            # except Exception as e:
                            #     print(f"Auto-commit failed: {str(e)}")

                            merged_count = sum(1 for nr_code in current_results if nr_code in [r[0] for r in data if len(r) >= 4 and r[3] in ['Yes', 'Maybe']])

                            messagebox.showinfo("Smart Match Complete",
                                f"✅ Consensus matching completed!\n\n"
                                f"📊 Loaded consensus results from 3 runs\n"
                                f"🔗 Merged {merged_count} previous results with new findings\n"
                                f"📋 Combined results shown with '/' separator\n"
                                f"💾 Auto-saved snapshot as ticksheet_working.csv\n"
                                f"🧠 High-confidence matches added to learning dictionary")
                        except Exception as e:
                            status_var.set("❌ Error loading results")
                            messagebox.showerror("Error", f"Failed to load consensus results: {e}")
                    else:
                        messagebox.showerror("File Not Found", "Consensus match output not found.")
                else:
                    status_var.set("❌ Consensus match failed")
                    error_msg = proc.stderr or proc.stdout or "Unknown error"
                    messagebox.showerror("Error", f"Consensus matcher failed for model '{model}':\n{error_msg}")
                    print(f"Consensus matcher failed. Error: {error_msg}")  # Debug output

            except Exception as e:
                if progress_data:
                    update_progress(progress_data, 0, "❌ Error", str(e))
                status_var.set("❌ Failed")
                messagebox.showerror("Error", str(e))
            finally:
                if progress_data:
                    close_progress(progress_data)
                # consensus_btn.config(state=tk.NORMAL)  # Button commented out

        threading.Thread(target=worker, daemon=True).start()

    def open_feature_dictionary():
        """Open feature dictionary search window with editing capabilities"""
        try:
            dict_window = tk.Toplevel(win)
            dict_window.title("📚 Feature Dictionary - Search, Edit & Manage")

            # Get screen dimensions for proper sizing
            screen_width = dict_window.winfo_screenwidth()
            screen_height = dict_window.winfo_screenheight()

            # Start with a good default size that fits most screens
            dict_width = min(1000, int(screen_width * 0.65))  # 65% of screen width, max 1000px
            dict_height = min(700, int(screen_height * 0.75))  # 75% of screen height, max 700px

            # Center the window
            dict_x = (screen_width - dict_width) // 2
            dict_y = (screen_height - dict_height) // 2

            # Ensure window doesn't go off screen
            if dict_x < 20:
                dict_x = 20
            if dict_y < 20:
                dict_y = 20

            dict_window.geometry(f"{dict_width}x{dict_height}+{dict_x}+{dict_y}")
            dict_window.configure(bg=COLORS['bg_primary'])

            # Make resizable with good min/max sizes
            dict_window.resizable(True, True)
            dict_window.minsize(800, 500)  # Minimum useful size
            dict_window.maxsize(screen_width-40, screen_height-80)  # Leave some margin

            # Add window size controls frame at the top
            size_controls = tk.Frame(dict_window, bg=COLORS['bg_secondary'], relief="flat", bd=1)
            size_controls.pack(fill="x", padx=5, pady=5)

            # Window size buttons with compact style
            size_btn_style = {
                "font": ("Arial", 8),
                "relief": "flat",
                "bd": 0,
                "pady": 2,
                "padx": 6,
                "cursor": "hand2"
            }

            tk.Label(size_controls, text="Window Size:", font=FONTS['small'],
                    bg=COLORS['bg_secondary'], fg=COLORS['text_primary']).pack(side="left", padx=(5,10))

            def resize_small():
                dict_window.geometry(f"{min(800, screen_width-100)}x{min(550, screen_height-150)}+{dict_x}+{dict_y}")

            def resize_medium():
                dict_window.geometry(f"{min(1000, screen_width-50)}x{min(700, screen_height-100)}+{dict_x}+{dict_y}")

            def resize_large():
                dict_window.geometry(f"{min(1200, screen_width-20)}x{min(800, screen_height-50)}+{dict_x}+{dict_y}")

            def resize_fullscreen():
                dict_window.state('zoomed')  # Maximize window

            small_btn = tk.Button(size_controls, text="Small", command=resize_small,
                                 bg="#64748B", fg="white", **size_btn_style)
            small_btn.pack(side="left", padx=2)

            medium_btn = tk.Button(size_controls, text="Medium", command=resize_medium,
                                  bg="#3B82F6", fg="white", **size_btn_style)
            medium_btn.pack(side="left", padx=2)

            large_btn = tk.Button(size_controls, text="Large", command=resize_large,
                                 bg="#059669", fg="white", **size_btn_style)
            large_btn.pack(side="left", padx=2)

            maximize_btn = tk.Button(size_controls, text="Maximize", command=resize_fullscreen,
                                    bg="#7C3AED", fg="white", **size_btn_style)
            maximize_btn.pack(side="left", padx=2)

            # Keep dictionary window on top and prevent it from going behind main UI
            dict_window.transient(win)
            dict_window.lift()
            dict_window.focus_force()

            # Initialize dictionary
            fd = FeatureDictionary()

            # Global variables for this window
            current_results = []

            # Results frame (create early so functions can reference it)
            results_frame = tk.Frame(dict_window)
            results_frame.pack(fill="both", expand=True, padx=15, pady=8)

            # Results tree (create early) with enhanced styling
            columns = ("ID", "Text", "Matched Feature", "Status", "Confidence", "Model", "Date", "Verified")

            # Create a style for the treeview
            style = ttk.Style()
            style.theme_use('clam')

            # Configure treeview colors for better readability
            style.configure("Dictionary.Treeview",
                          background=COLORS['row_even'],
                          foreground=COLORS['text_primary'],
                          fieldbackground=COLORS['row_even'],
                          font=FONTS['body'],
                          rowheight=28)  # Increased row height for better readability

            style.configure("Dictionary.Treeview.Heading",
                          background=COLORS['bg_accent'],
                          foreground=COLORS['text_primary'],
                          font=FONTS['subheading'])

            # Configure alternating row colors
            style.map("Dictionary.Treeview",
                     background=[('selected', COLORS['blue_primary']),
                               ('!selected', COLORS['row_even'])])

            results_tree = ttk.Treeview(results_frame, columns=columns, show="headings",
                                      height=20, style="Dictionary.Treeview")

            # Configure columns with better spacing and clickability info
            results_tree.heading("ID", text="ID")
            results_tree.heading("Text", text="Original Text")
            results_tree.heading("Matched Feature", text="Matched Feature (Hover for Full)")
            results_tree.heading("Status", text="Status (Click)")
            results_tree.heading("Confidence", text="Confidence (Click)")
            results_tree.heading("Model", text="Model")
            results_tree.heading("Date", text="Date")
            results_tree.heading("Verified", text="✓")

            # Column widths optimized for readability
            results_tree.column("ID", width=0, stretch=False)  # Hidden
            results_tree.column("Text", width=280, minwidth=180)
            results_tree.column("Matched Feature", width=220, minwidth=140)
            results_tree.column("Status", width=140, minwidth=120)
            results_tree.column("Confidence", width=90, minwidth=80)
            results_tree.column("Model", width=90, minwidth=70)
            results_tree.column("Date", width=90, minwidth=80)
            results_tree.column("Verified", width=50, minwidth=40)

            # Hide ID column header
            results_tree.heading("ID", text="")

            # Scrollbars
            v_scroll = ttk.Scrollbar(results_frame, orient="vertical", command=results_tree.yview)
            h_scroll = ttk.Scrollbar(results_frame, orient="horizontal", command=results_tree.xview)
            results_tree.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)

            # Pack tree and scrollbars
            results_tree.grid(row=0, column=0, sticky="nsew")
            v_scroll.grid(row=0, column=1, sticky="ns")
            h_scroll.grid(row=1, column=0, sticky="ew")

            results_frame.grid_rowconfigure(0, weight=1)
            results_frame.grid_columnconfigure(0, weight=1)

            # Status bar with enhanced styling
            status_frame = tk.Frame(dict_window, bg=COLORS['bg_accent'], relief="flat", bd=1)
            status_frame.pack(fill="x")

            status_label = tk.Label(status_frame, text="Ready to search",
                                   font=FONTS['body'], bg=COLORS['bg_accent'], fg=COLORS['text_primary'])
            status_label.pack(side="left", padx=15, pady=8)

            def safe_execute(func, *args, **kwargs):
                """Safely execute a function with error handling"""
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    print(f"Error in {func.__name__}: {str(e)}")
                    messagebox.showerror("Error", f"An error occurred: {str(e)}")
                    return None

            def update_results_view(results):
                """Update the results tree with new data - optimized for speed with alternating colors"""
                try:
                    nonlocal current_results
                    current_results = results

                    # Clear previous results
                    for item in results_tree.get_children():
                        results_tree.delete(item)

                    # Limit results for UI responsiveness
                    display_results = results[:200] if len(results) > 200 else results

                    # Configure alternating row colors
                    results_tree.tag_configure('even', background=COLORS['row_even'])
                    results_tree.tag_configure('odd', background=COLORS['row_odd'])

                    # Insert items with alternating colors
                    for i, result in enumerate(display_results):
                        confidence_pct = f"{result['confidence']:.0%}" if result.get('confidence') else "N/A"
                        verified_icon = "✅" if result.get('user_verified') else ""

                        # Color coding for status - clean display without [Click]
                        status_text = result.get('status', 'Unknown')
                        if status_text == 'Yes':
                            status_color = "✅ Yes"
                        elif status_text == 'No':
                            status_color = "❌ No"
                        elif status_text == 'Maybe':
                            status_color = "❓ Maybe"
                        else:
                            status_color = status_text

                        # Make confidence clickable too
                        confidence_pct = f"{result['confidence']:.0%}" if result.get('confidence') else "N/A"

                        # Truncate long text for faster display with better spacing
                        text_display = (result.get('text', '')[:45] + "...") if len(result.get('text', '')) > 45 else result.get('text', '')
                        name_display = (result.get('name', '')[:28] + "...") if len(result.get('name', '')) > 28 else result.get('name', '')

                        # Determine row tag for alternating colors
                        row_tag = 'even' if i % 2 == 0 else 'odd'

                        item = results_tree.insert("", "end", values=(
                            result.get('id', ''),  # Hidden but stored for reference
                            text_display,
                            name_display,
                            status_color,
                            confidence_pct,
                            result.get('model', ''),
                            result.get('timestamp', '')[:10],  # Just date
                            verified_icon
                        ), tags=(row_tag,))

                    # Update status with better formatting
                    if len(results) > 200:
                        status_label.config(text=f"📊 Showing first 200 of {len(results)} results (filtered for performance)")
                    else:
                        status_label.config(text=f"📊 Showing {len(results)} results")

                except Exception as e:
                    print(f"Error updating results view: {str(e)}")
                    messagebox.showerror("Display Error", f"Failed to update results: {str(e)}")

            def refresh_results():
                """Refresh the current view"""
                try:
                    query = search_var.get().strip() if 'search_var' in locals() else ""
                    if query:
                        perform_search()
                    else:
                        list_all_features()
                except Exception as e:
                    print(f"Error refreshing results: {str(e)}")
                    status_label.config(text="Error refreshing results")

            def perform_search():
                try:
                    query = search_var.get().strip()
                    if not query:
                        return

                    status_label.config(text="Searching...")
                    dict_window.update_idletasks()

                    results = fd.search_features(query, limit=1000)
                    update_results_view(results)
                    status_label.config(text=f"Found {len(results)} features matching '{query}'")
                except Exception as e:
                    print(f"Error in search: {str(e)}")
                    status_label.config(text="Search failed")
                    messagebox.showerror("Search Error", f"Search failed: {str(e)}")

            def list_all_features():
                """List all features in the dictionary"""
                try:
                    status_label.config(text="Loading all features...")
                    dict_window.update_idletasks()

                    all_results = fd.search_features("", limit=1000)  # Empty query returns all
                    update_results_view(all_results)
                    status_label.config(text=f"Showing {min(len(all_results), 500)} of {len(all_results)} total features")
                except Exception as e:
                    print(f"Error listing features: {str(e)}")
                    status_label.config(text="Failed to load features")
                    messagebox.showerror("Load Error", f"Failed to load features: {str(e)}")

            def list_verified_features():
                """List only user-verified features"""
                try:
                    status_label.config(text="Loading verified features...")
                    dict_window.update_idletasks()

                    all_results = fd.search_features("", limit=1000)  # Get all first
                    verified_results = [f for f in all_results if f.get('user_verified', False)]

                    update_results_view(verified_results)
                    status_label.config(text=f"Showing {len(verified_results)} user-verified features")
                except Exception as e:
                    print(f"Error listing verified features: {str(e)}")
                    status_label.config(text="Failed to load verified features")
                    messagebox.showerror("Load Error", f"Failed to load verified features: {str(e)}")

            def on_tree_click(event):
                """Handle clicks on the tree for status and confidence columns"""
                try:
                    region = results_tree.identify_region(event.x, event.y)
                    if region == "cell":
                        column = results_tree.identify_column(event.x)
                        item = results_tree.identify_row(event.y)
                        if item:
                            if column == "#4":  # Status column (4th column)
                                toggle_status(item)
                            elif column == "#5":  # Confidence column (5th column)
                                toggle_confidence(item)
                except Exception as e:
                    print(f"Error in tree click: {str(e)}")

            def on_tree_motion(event):
                """Handle mouse motion for tooltips"""
                try:
                    region = results_tree.identify_region(event.x, event.y)
                    if region == "cell":
                        column = results_tree.identify_column(event.x)
                        item = results_tree.identify_row(event.y)
                        if item and column == "#3":  # Matched Feature column (3rd column)
                            show_feature_tooltip(event, item)
                        else:
                            hide_tooltip()
                    else:
                        hide_tooltip()
                except Exception as e:
                    pass  # Silently handle tooltip errors

            # Tooltip variables
            tooltip_window = None

            def show_feature_tooltip(event, item):
                """Show tooltip with full feature text"""
                try:
                    nonlocal tooltip_window
                    hide_tooltip()  # Hide any existing tooltip

                    values = results_tree.item(item)['values']
                    if not values or len(values) < 3:
                        return

                    feature_id = values[0]
                    if not feature_id:
                        return

                    # Get full feature details
                    feature = fd.get_feature_by_id(feature_id)
                    if not feature:
                        return

                    full_text = feature.get('name', 'No feature name available')

                    # Create tooltip window
                    tooltip_window = tk.Toplevel(dict_window)
                    tooltip_window.wm_overrideredirect(True)
                    tooltip_window.configure(bg="#FFFBEB", relief="solid", bd=1)

                    # Position tooltip near cursor
                    x = event.x_root + 10
                    y = event.y_root + 10

                    # Ensure tooltip stays on screen
                    screen_width = tooltip_window.winfo_screenwidth()
                    screen_height = tooltip_window.winfo_screenheight()

                    # Create label with word wrapping
                    label = tk.Label(tooltip_window, text=full_text,
                                   bg="#FFFBEB", fg="#1F2937",
                                   font=FONTS['small'],
                                   wraplength=300, justify="left",
                                   padx=8, pady=6)
                    label.pack()

                    # Update tooltip window to get actual size
                    tooltip_window.update_idletasks()
                    tooltip_width = tooltip_window.winfo_reqwidth()
                    tooltip_height = tooltip_window.winfo_reqheight()

                    # Adjust position if tooltip would go off screen
                    if x + tooltip_width > screen_width:
                        x = screen_width - tooltip_width - 10
                    if y + tooltip_height > screen_height:
                        y = event.y_root - tooltip_height - 10

                    tooltip_window.geometry(f"+{x}+{y}")

                except Exception as e:
                    hide_tooltip()

            def hide_tooltip():
                """Hide the tooltip"""
                try:
                    nonlocal tooltip_window
                    if tooltip_window:
                        tooltip_window.destroy()
                        tooltip_window = None
                except:
                    tooltip_window = None

            def toggle_status(item):
                """Quickly toggle status between Yes/No/Maybe"""
                try:
                    values = results_tree.item(item)['values']
                    if not values or len(values) < 1:
                        return

                    feature_id = values[0]
                    current_status = values[3]

                    # Cycle through statuses: Yes -> Maybe -> No -> Yes
                    if "Yes" in current_status:
                        new_status = "Maybe"
                        new_display = "❓ Maybe"
                    elif "Maybe" in current_status:
                        new_status = "No"
                        new_display = "❌ No"
                    else:  # "No" or anything else
                        new_status = "Yes"
                        new_display = "✅ Yes"

                    # Update in database
                    if fd.update_feature(feature_id, match_status=new_status):
                        # Mark as user verified
                        fd.mark_as_verified(feature_id, True, f"User quick-changed to {new_status}")

                        # Update display
                        new_values = list(values)
                        new_values[3] = new_display
                        new_values[7] = "✅"  # Mark as verified
                        results_tree.item(item, values=new_values)

                        status_label.config(text=f"Status changed to {new_status}")
                    else:
                        messagebox.showerror("Error", "Failed to update status.")

                except Exception as e:
                    print(f"Error toggling status: {str(e)}")
                    messagebox.showerror("Status Error", f"Failed to change status: {str(e)}")

            def toggle_confidence(item):
                """Quickly cycle through confidence levels"""
                try:
                    values = results_tree.item(item)['values']
                    if not values or len(values) < 5:
                        return

                    feature_id = values[0]
                    current_confidence = values[4]

                    # Parse current confidence percentage
                    try:
                        current_pct = int(current_confidence.replace('%', '').replace('N/A', '0'))
                    except:
                        current_pct = 0

                    # Cycle through confidence levels: 0% -> 25% -> 50% -> 75% -> 90% -> 100% -> 0%
                    confidence_levels = [0, 25, 50, 75, 90, 100]
                    try:
                        current_index = confidence_levels.index(current_pct)
                        next_index = (current_index + 1) % len(confidence_levels)
                    except ValueError:
                        next_index = 1  # Default to 25% if current value not in list

                    new_confidence_pct = confidence_levels[next_index]
                    new_confidence = new_confidence_pct / 100.0

                    # Update in database
                    if fd.update_feature(feature_id, confidence=new_confidence):
                        # Mark as user verified
                        fd.mark_as_verified(feature_id, True, f"User changed confidence to {new_confidence_pct}%")

                        # Update display
                        new_values = list(values)
                        new_values[4] = f"{new_confidence_pct}%"
                        new_values[7] = "✅"  # Mark as verified
                        results_tree.item(item, values=new_values)

                        status_label.config(text=f"Confidence changed to {new_confidence_pct}%")
                    else:
                        messagebox.showerror("Error", "Failed to update confidence.")

                except Exception as e:
                    print(f"Error toggling confidence: {str(e)}")
                    messagebox.showerror("Confidence Error", f"Failed to change confidence: {str(e)}")

            def edit_selected():
                """Edit the selected feature"""
                try:
                    selection = results_tree.selection()
                    if not selection:
                        messagebox.showwarning("No Selection", "Please select a feature to edit.")
                        return

                    item = results_tree.item(selection[0])
                    if not item['values'] or len(item['values']) < 1:
                        messagebox.showerror("Error", "Invalid selection.")
                        return

                    feature_id = item['values'][0]
                    if not feature_id:
                        messagebox.showerror("Error", "No feature ID found.")
                        return

                    # Get full feature details
                    feature = fd.get_feature_by_id(feature_id)
                    if not feature:
                        messagebox.showerror("Error", "Could not load feature details.")
                        return

                    open_edit_dialog(feature, refresh_results)
                except Exception as e:
                    print(f"Error editing feature: {str(e)}")
                    messagebox.showerror("Edit Error", f"Failed to edit feature: {str(e)}")

            def delete_selected():
                """Delete the selected feature"""
                try:
                    selection = results_tree.selection()
                    if not selection:
                        messagebox.showwarning("No Selection", "Please select a feature to delete.")
                        return

                    item = results_tree.item(selection[0])
                    feature_id = item['values'][0]
                    feature_text = item['values'][1]

                    # Confirm deletion
                    if messagebox.askyesno("Confirm Deletion",
                                         f"Are you sure you want to delete this feature?\n\n'{feature_text}'\n\nThis action cannot be undone."):
                        if fd.delete_feature(feature_id):
                            messagebox.showinfo("Deleted", "Feature deleted successfully.")
                            refresh_results()
                        else:
                            messagebox.showerror("Error", "Failed to delete feature.")
                except Exception as e:
                    print(f"Error deleting feature: {str(e)}")
                    messagebox.showerror("Delete Error", f"Failed to delete feature: {str(e)}")

            def mark_as_false_flag():
                """Mark selected feature as false flag (incorrect match)"""
                try:
                    selection = results_tree.selection()
                    if not selection:
                        messagebox.showwarning("No Selection", "Please select a feature to mark as false flag.")
                        return

                    item = results_tree.item(selection[0])
                    feature_id = item['values'][0]

                    # Update to mark as false flag
                    if fd.update_feature(feature_id, match_status="No", confidence=0.0,
                                       reasoning="User marked as false flag - incorrect match"):
                        if fd.mark_as_verified(feature_id, True, "False flag - incorrect match"):
                            messagebox.showinfo("Updated", "Feature marked as false flag.")
                            refresh_results()
                        else:
                            messagebox.showwarning("Partial Update", "Feature updated but verification flag failed.")
                    else:
                        messagebox.showerror("Error", "Failed to mark as false flag.")
                except Exception as e:
                    print(f"Error marking false flag: {str(e)}")
                    messagebox.showerror("False Flag Error", f"Failed to mark as false flag: {str(e)}")

            def mark_as_correct_match():
                """Mark selected feature as correct match (untick false flag)"""
                try:
                    selection = results_tree.selection()
                    if not selection:
                        messagebox.showwarning("No Selection", "Please select a feature to mark as correct match.")
                        return

                    item = results_tree.item(selection[0])
                    feature_id = item['values'][0]

                    # Get current feature to preserve other details
                    feature = fd.get_feature_by_id(feature_id)
                    if not feature:
                        messagebox.showerror("Error", "Could not load feature details.")
                        return

                    # Update to mark as correct match with reasonable confidence
                    confidence = max(0.7, feature.get('confidence', 0.7))  # At least 70% confidence
                    if fd.update_feature(feature_id, match_status="Yes", confidence=confidence,
                                       reasoning="User verified as correct match"):
                        if fd.mark_as_verified(feature_id, True, "User verified as correct"):
                            messagebox.showinfo("Updated", "Feature marked as correct match.")
                            refresh_results()
                        else:
                            messagebox.showwarning("Partial Update", "Feature updated but verification flag failed.")
                    else:
                        messagebox.showerror("Error", "Failed to mark as correct match.")
                except Exception as e:
                    print(f"Error marking correct match: {str(e)}")
                    messagebox.showerror("Correct Match Error", f"Failed to mark as correct match: {str(e)}")

            def bulk_commit_changes():
                """Commit all user-verified changes to improve the dictionary database"""
                def worker():
                    try:
                        # Get all verified features
                        status_label.config(text="🔍 Finding verified changes...")
                        dict_window.update_idletasks()

                        verified_features = fd.search_features("", limit=10000)  # Get all
                        user_verified = [f for f in verified_features if f.get('user_verified', False)]

                        if not user_verified:
                            status_label.config(text="No changes to commit")
                            messagebox.showinfo("No Changes", "No user-verified changes found to commit.")
                            return

                        if not messagebox.askyesno("Commit Changes",
                                                 f"Commit {len(user_verified)} verified changes to improve the dictionary?\n\n" +
                                                 "This will:\n" +
                                                 "• Update confidence scores for similar future matches\n" +
                                                 "• Improve AI learning from your corrections\n" +
                                                 "• Make the system smarter over time"):
                            status_label.config(text="Commit cancelled")
                            return

                        status_label.config(text="💾 Committing changes to dictionary...")
                        dict_window.update_idletasks()

                        # Process each verified change to improve dictionary
                        committed_count = 0
                        total = len(user_verified)

                        for i, feature in enumerate(user_verified):
                            try:
                                # Update status for responsiveness
                                if i % 5 == 0:  # Update every 5 items
                                    status_label.config(text=f"💾 Committing {i+1}/{total}...")
                                    dict_window.update_idletasks()

                                # For correct matches (Yes), increase confidence pattern recognition
                                if feature.get('status') == 'Yes':
                                    # Mark this pattern as confirmed good
                                    fd.mark_as_verified(feature['id'], True,
                                                      f"User confirmed - High confidence pattern")
                                    committed_count += 1

                                # For false flags (No), help prevent similar future mistakes
                                elif feature.get('status') == 'No':
                                    # Mark this pattern as confirmed bad
                                    fd.mark_as_verified(feature['id'], True,
                                                      f"User rejected - Avoid this pattern")
                                    committed_count += 1

                            except Exception as e:
                                print(f"Error committing feature {feature.get('id')}: {str(e)}")
                                continue

                        status_label.config(text=f"✅ Committed {committed_count} changes!")
                        messagebox.showinfo("Success",
                                          f"Successfully committed {committed_count} changes!\n\n" +
                                          "The dictionary has been improved with your feedback.\n" +
                                          "Future analyses will be more accurate.")

                        refresh_results()

                    except Exception as e:
                        print(f"Error committing changes: {str(e)}")
                        status_label.config(text="❌ Commit failed")
                        messagebox.showerror("Commit Error", f"Failed to commit changes: {str(e)}")

                # Run in background thread for responsiveness
                import threading
                threading.Thread(target=worker, daemon=True).start()

            def clear_all_verifications():
                """Clear all user verification flags (for testing/reset)"""
                try:
                    if messagebox.askyesno("Confirm Reset",
                                         "Are you sure you want to clear ALL user verification flags?\n\n" +
                                         "This will not delete features, but will mark them as unverified.\n" +
                                         "This action cannot be undone."):

                        # This would require a new method in FeatureDictionary
                        # For now, just show a message
                        messagebox.showinfo("Feature Not Implemented",
                                          "This feature would clear all verification flags.\n" +
                                          "Currently not implemented for safety.")

                except Exception as e:
                    print(f"Error clearing verifications: {str(e)}")
                    messagebox.showerror("Clear Error", f"Failed to clear verifications: {str(e)}")

            # Search frame with enhanced styling
            search_frame = tk.Frame(dict_window, bg=COLORS['bg_secondary'], pady=12)
            search_frame.pack(fill="x", padx=15, pady=8, before=results_frame)

            # Instruction label with better styling
            instruction_frame = tk.Frame(search_frame, bg=COLORS['bg_secondary'])
            instruction_frame.pack(fill="x", pady=(0, 8))

            instruction_text = "💡 Click Status to toggle • 📝 Edit for full details • Window is resizable"
            instruction_label = tk.Label(instruction_frame, text=instruction_text,
                                       font=FONTS['small'], fg=COLORS['text_secondary'], bg=COLORS['bg_secondary'],
                                       wraplength=500, justify="left")
            instruction_label.pack(anchor="w")

            # Search controls with enhanced styling - reorganized for single row
            search_controls = tk.Frame(search_frame, bg=COLORS['bg_secondary'])
            search_controls.pack(fill="x", pady=5)

            # Left side - search input
            search_left = tk.Frame(search_controls, bg=COLORS['bg_secondary'])
            search_left.pack(side="left", fill="x")

            tk.Label(search_left, text="🔍 Search:", font=FONTS['subheading'],
                    bg=COLORS['bg_secondary'], fg=COLORS['text_primary']).pack(side="left")

            search_var = tk.StringVar()
            search_entry = tk.Entry(search_left, textvariable=search_var, font=FONTS['body'],
                                  width=25, relief="flat", bd=2)
            search_entry.pack(side="left", padx=(8, 10))

            # Enhanced button styling for dictionary - smaller and more compact
            dict_btn_style = {
                "font": FONTS['small'],  # Use smaller font
                "relief": "flat",
                "bd": 0,
                "pady": 3,   # Even more reduced padding
                "padx": 5,   # Reduced padding
                "cursor": "hand2"
            }

            search_btn = tk.Button(search_left, text="Search", command=lambda: safe_execute(perform_search),
                                  bg=COLORS['blue_primary'], fg="white", **dict_btn_style)
            search_btn.pack(side="left", padx=2)

            # Right side - ALL ACTION BUTTONS IN ONE ROW
            buttons_right = tk.Frame(search_controls, bg=COLORS['bg_secondary'])
            buttons_right.pack(side="right", padx=5)

            # View buttons
            list_all_btn = tk.Button(buttons_right, text="📋 All", command=lambda: safe_execute(list_all_features),
                                    bg=COLORS['green_primary'], fg="white", **dict_btn_style)
            list_all_btn.pack(side="left", padx=1)

            verified_btn = tk.Button(buttons_right, text="✅ Verified", command=lambda: safe_execute(list_verified_features),
                                    bg="#059669", fg="white", **dict_btn_style)
            verified_btn.pack(side="left", padx=1)

            # Action buttons - all in one row
            edit_details_btn = tk.Button(buttons_right, text="📝 Edit", command=lambda: safe_execute(edit_selected),
                                        bg="#8B5CF6", fg="white", **dict_btn_style)
            edit_details_btn.pack(side="left", padx=1)

            delete_btn = tk.Button(buttons_right, text="🗑️ Del", command=lambda: safe_execute(delete_selected),
                                  bg=COLORS['red_primary'], fg="white", **dict_btn_style)
            delete_btn.pack(side="left", padx=1)

            false_flag_btn = tk.Button(buttons_right, text="❌ Flag", command=lambda: safe_execute(mark_as_false_flag),
                                      bg=COLORS['orange_primary'], fg="white", **dict_btn_style)
            false_flag_btn.pack(side="left", padx=1)

            correct_btn = tk.Button(buttons_right, text="✅ OK", command=lambda: safe_execute(mark_as_correct_match),
                                   bg="#10B981", fg="white", **dict_btn_style)
            correct_btn.pack(side="left", padx=1)

            commit_btn = tk.Button(buttons_right, text="💾 Save", command=lambda: safe_execute(bulk_commit_changes),
                                  bg=COLORS['blue_primary'], fg="white", **dict_btn_style)
            commit_btn.pack(side="left", padx=1)

            # Bind events to tree
            results_tree.bind("<Button-1>", on_tree_click)  # Click for status/confidence toggle
            results_tree.bind("<Motion>", on_tree_motion)    # Motion for tooltips
            results_tree.bind("<Leave>", lambda e: hide_tooltip())  # Hide tooltip when leaving tree

            # Bind Enter key to search
            search_entry.bind("<Return>", lambda e: safe_execute(perform_search))

            # Load initial data safely
            try:
                list_all_features()
            except Exception as e:
                print(f"Error loading initial data: {str(e)}")
                status_label.config(text="Failed to load initial data")

            # Focus on search entry
            search_entry.focus()

            # Handle window closing
            def on_closing():
                try:
                    dict_window.destroy()
                except:
                    pass

            dict_window.protocol("WM_DELETE_WINDOW", on_closing)

        except Exception as e:
            print(f"Error opening dictionary: {str(e)}")
            messagebox.showerror("Dictionary Error", f"Failed to open dictionary: {str(e)}")
            try:
                dict_window.destroy()
            except:
                pass

    def open_edit_dialog(feature, refresh_callback):
        """Open edit dialog for a feature"""
        try:
            edit_window = tk.Toplevel(win)
            edit_window.title(f"✏️ Edit Feature #{feature['id']}")

            # Position edit dialog in center-left of screen
            screen_width = edit_window.winfo_screenwidth()
            screen_height = edit_window.winfo_screenheight()

            dialog_width = 600
            dialog_height = 500
            dialog_x = int(screen_width * 0.2)  # 20% from left
            dialog_y = int((screen_height - dialog_height) / 2)  # Centered vertically

            edit_window.geometry(f"{dialog_width}x{dialog_height}+{dialog_x}+{dialog_y}")
            edit_window.configure(bg=COLORS['bg_primary'])

            # Keep edit dialog properly managed
            edit_window.transient(win)
            edit_window.grab_set()  # Modal dialog
            edit_window.focus_set()
            edit_window.lift()

            # Main frame with enhanced styling
            main_frame = tk.Frame(edit_window, bg=COLORS['bg_primary'], padx=25, pady=25)
            main_frame.pack(fill="both", expand=True)

            # Original text (read-only) with better styling
            tk.Label(main_frame, text="Original Text:", font=FONTS['subheading'],
                    bg=COLORS['bg_primary'], fg=COLORS['text_primary']).pack(anchor="w")

            text_frame = tk.Frame(main_frame, bg=COLORS['bg_primary'])
            text_frame.pack(fill="x", pady=(8, 18))

            text_display = tk.Text(text_frame, height=3, font=FONTS['body'],
                                  state="disabled", wrap="word", bg=COLORS['bg_secondary'],
                                  relief="flat", bd=2)
            text_display.pack(fill="x")
            text_display.config(state="normal")
            text_display.insert("1.0", feature.get('text', ''))
            text_display.config(state="disabled")

            # Matched Code with enhanced styling
            tk.Label(main_frame, text="Matched Code:", font=FONTS['subheading'],
                    bg=COLORS['bg_primary'], fg=COLORS['text_primary']).pack(anchor="w")

            code_var = tk.StringVar(value=feature.get('code', '') or "")
            code_entry = tk.Entry(main_frame, textvariable=code_var, font=FONTS['body'],
                                 relief="flat", bd=2, bg="#FFFFFF")
            code_entry.pack(fill="x", pady=(8, 18))

            # Matched Name with enhanced styling
            tk.Label(main_frame, text="Matched Feature Name:", font=FONTS['subheading'],
                    bg=COLORS['bg_primary'], fg=COLORS['text_primary']).pack(anchor="w")

            name_var = tk.StringVar(value=feature.get('name', '') or "")
            name_entry = tk.Entry(main_frame, textvariable=name_var, font=FONTS['body'],
                                 relief="flat", bd=2, bg="#FFFFFF")
            name_entry.pack(fill="x", pady=(8, 18))

            # Status with enhanced styling
            tk.Label(main_frame, text="Match Status:", font=FONTS['subheading'],
                    bg=COLORS['bg_primary'], fg=COLORS['text_primary']).pack(anchor="w")

            status_var = tk.StringVar(value=feature.get('status', 'No') or "No")
            status_frame = tk.Frame(main_frame, bg=COLORS['bg_primary'])
            status_frame.pack(fill="x", pady=(8, 18))

            for status in ["Yes", "No", "Maybe"]:
                rb = tk.Radiobutton(status_frame, text=status, variable=status_var,
                                  value=status, font=FONTS['body'], bg=COLORS['bg_primary'],
                                  fg=COLORS['text_primary'], selectcolor="#FFFFFF")
                rb.pack(side="left", padx=(0, 20))

            # Confidence with enhanced styling
            tk.Label(main_frame, text="Confidence (%):", font=FONTS['subheading'],
                    bg=COLORS['bg_primary'], fg=COLORS['text_primary']).pack(anchor="w")

            confidence_val = feature.get('confidence', 0) or 0
            confidence_var = tk.StringVar(value=str(int(confidence_val * 100)))
            confidence_entry = tk.Entry(main_frame, textvariable=confidence_var, font=FONTS['body'],
                                       relief="flat", bd=2, bg="#FFFFFF")
            confidence_entry.pack(fill="x", pady=(8, 18))

            # Reasoning with enhanced styling
            tk.Label(main_frame, text="Reasoning:", font=FONTS['subheading'],
                    bg=COLORS['bg_primary'], fg=COLORS['text_primary']).pack(anchor="w")

            reasoning_frame = tk.Frame(main_frame, bg=COLORS['bg_primary'])
            reasoning_frame.pack(fill="both", expand=True, pady=(8, 18))

            reasoning_text = tk.Text(reasoning_frame, height=4, font=FONTS['body'], wrap="word",
                                   relief="flat", bd=2, bg="#FFFFFF")
            reasoning_scroll = ttk.Scrollbar(reasoning_frame, orient="vertical", command=reasoning_text.yview)
            reasoning_text.configure(yscrollcommand=reasoning_scroll.set)

            reasoning_text.pack(side="left", fill="both", expand=True)
            reasoning_scroll.pack(side="right", fill="y")

            reasoning_text.insert("1.0", feature.get('reasoning', '') or "")

            def save_changes():
                try:
                    # Get values
                    new_code = code_var.get().strip()
                    new_name = name_var.get().strip()
                    new_status = status_var.get()

                    # Validate confidence
                    try:
                        conf_val = confidence_var.get().strip()
                        new_confidence = float(conf_val) / 100.0 if conf_val else 0.0
                        if new_confidence < 0 or new_confidence > 1:
                            raise ValueError("Confidence must be between 0 and 100")
                    except ValueError:
                        messagebox.showerror("Invalid Input", "Confidence must be a number between 0 and 100.")
                        return

                    new_reasoning = reasoning_text.get("1.0", "end-1c").strip()

                    # Initialize dictionary
                    fd = FeatureDictionary()

                    # Update feature
                    success = fd.update_feature(
                        feature['id'],
                        matched_code=new_code,
                        matched_name=new_name,
                        match_status=new_status,
                        confidence=new_confidence,
                        reasoning=new_reasoning
                    )

                    if success:
                        # Mark as user verified
                        fd.mark_as_verified(feature['id'], True, "User edited")
                        messagebox.showinfo("Success", "Feature updated successfully!")
                        edit_window.destroy()
                        if refresh_callback:
                            refresh_callback()
                    else:
                        messagebox.showerror("Error", "Failed to update feature.")

                except Exception as e:
                    print(f"Error saving changes: {str(e)}")
                    messagebox.showerror("Error", f"An error occurred: {str(e)}")

            def cancel_edit():
                try:
                    edit_window.destroy()
                except:
                    pass

            # Buttons with enhanced styling
            button_frame = tk.Frame(main_frame, bg=COLORS['bg_primary'])
            button_frame.pack(fill="x", pady=(15, 0))

            # Enhanced button styling for dialogs
            dialog_btn_style = {
                "font": FONTS['button'],
                "relief": "flat",
                "bd": 0,
                "pady": 12,
                "padx": 20,
                "cursor": "hand2"
            }

            save_btn = tk.Button(button_frame, text="💾 Save Changes", command=save_changes,
                                bg=COLORS['blue_primary'], fg="white", **dialog_btn_style)
            save_btn.pack(side="right", padx=(12, 0))

            cancel_btn = tk.Button(button_frame, text="❌ Cancel", command=cancel_edit,
                                  bg=COLORS['text_secondary'], fg="white", **dialog_btn_style)
            cancel_btn.pack(side="right")

            # Handle window closing
            edit_window.protocol("WM_DELETE_WINDOW", cancel_edit)

            # Focus on first editable field
            code_entry.focus()

        except Exception as e:
            print(f"Error opening edit dialog: {str(e)}")
            messagebox.showerror("Edit Dialog Error", f"Failed to open edit dialog: {str(e)}")
            try:
                edit_window.destroy()
            except:
                pass

    def on_open_exports():
        folder = model_exports_dir(model_var.get())
        if sys.platform.startswith("win"): subprocess.Popen(["explorer", str(folder)])
        elif sys.platform == "darwin": subprocess.Popen(["open", str(folder)])
        else: subprocess.Popen(["xdg-open", str(folder)])
    # Enhanced button styling with modern design - aligned with contemporary aesthetics
    btn_style = {
        "font": FONTS['button'],
        "relief": "flat",
        "bd": 0,
        "pady": 8,  # Modern padding
        "padx": 16, # More generous horizontal padding
        "cursor": "hand2",
        "activebackground": "#1E40AF",  # Modern hover effect
        "activeforeground": "#FFFFFF"
    }

    # Load buttons - Modern gray theme for secondary actions
    load_txt_btn = tk.Button(btn_row, text="📄 Load .txt", command=on_load_file,
                            bg=COLORS['bg_accent'], fg=COLORS['text_primary'], **btn_style)
    load_txt_btn.pack(side="left", padx=(0, 4))

    load_csv_btn = tk.Button(btn_row, text="📊 Load CSV", command=lambda: on_load_dual_csv(),
                            bg=COLORS['bg_accent'], fg=COLORS['text_primary'], **btn_style)
    load_csv_btn.pack(side="left", padx=4)

    # AI buttons - Professional blue theme for primary actions
    # REMOVED: Intelligent button - Single-language analysis, superseded by Learning MEGA
    # intelligent_btn = tk.Button(btn_row, text="🧠 Intelligent", width=12, command=lambda: on_run_intelligent(),
    #                            bg=COLORS['blue_primary'], fg="white", **btn_style)
    # intelligent_btn.pack(side="left", padx=4)

    # REMOVED: Smart (3x) button - Consensus-only, redundant with Learning MEGA's enhanced consensus
    # consensus_btn = tk.Button(btn_row, text="✨ Smart (3x)", width=12, command=lambda: on_run_consensus(),
    #                          bg=COLORS['blue_secondary'], fg="white", **btn_style)
    # consensus_btn.pack(side="left", padx=4)

    # REMOVED: Sequential MEGA button - Dual-language but without learning enhancement, superseded by Learning MEGA
    # mega_btn = tk.Button(btn_row, text="🚀 MEGA", width=10, command=lambda: on_run_mega(),
    #                     bg=COLORS['blue_primary'], fg="white", **btn_style)
    # mega_btn.pack(side="left", padx=4)

    # 🎯 CANONICAL WORKFLOW: Learning Sequential MEGA button - RECOMMENDED PRIMARY ANALYSIS METHOD
    # This is the main analysis workflow that should be used for all feature matching tasks.
    # Benefits: Enhanced with learned dictionary knowledge, dual-language support, learning hints,
    #           automatic knowledge expansion, and improved accuracy over time.
    # Use this instead of other analysis buttons for best results.
    learning_mega_btn = tk.Button(btn_row, text="🎓 Learning", width=10, command=lambda: on_run_learning_mega(),
                                 bg=COLORS['blue_secondary'], fg="white", **btn_style)
    learning_mega_btn.pack(side="left", padx=4)

    # Utility buttons - Modern professional styling
    save_btn = tk.Button(btn_row, text="💾 Save", command=lambda: save_snapshot(model_var.get(), table, status_var),
                        bg=COLORS['text_secondary'], fg="white", **btn_style)
    save_btn.pack(side="left", padx=4)

    export_btn = tk.Button(btn_row, text="📁 Exports", command=on_open_exports,
                          bg=COLORS['text_secondary'], fg="white", **btn_style)
    export_btn.pack(side="left", padx=4)

    # Dictionary button - Modern accent
    dict_btn = tk.Button(btn_row, text="📚 Dictionary", command=lambda: open_feature_dictionary(),
                        bg=COLORS['blue_primary'], fg="white", **btn_style)
    dict_btn.pack(side="left", padx=4)

    # Modern info label with clean contemporary design
    info_frame = tk.Frame(mid, bg=COLORS['bg_accent'], relief="flat", bd=0)
    info_frame.pack(fill="x", pady=(8, 0))

    info_text = "💡 Quick Guide: 🎓 Learning: PRIMARY workflow with enhanced AI knowledge •  Dictionary: Search & edit features • Press F11 for fullscreen, ESC to exit"
    info_label = tk.Label(info_frame, text=info_text, font=FONTS['small'],
                         fg=COLORS['text_secondary'], bg=COLORS['bg_accent'],
                         wraplength=1400, justify="left")
    info_label.pack(anchor="w", padx=12, pady=6)

    # Modern table frame with clean professional design
    table_frame = tk.Frame(win, bg=COLORS['bg_primary'], relief="flat", bd=0)
    table_frame.pack(fill="both", expand=True, padx=15, pady=(0, 15))

    # Modern header for the table area
    table_header = tk.Frame(table_frame, bg=COLORS['bg_accent'], height=40)
    table_header.pack(fill="x")
    table_header.pack_propagate(False)

    table_title = tk.Label(table_header, text="📋 Feature Analysis Results",
                          font=FONTS['heading'], bg=COLORS['bg_accent'], fg=COLORS['text_primary'])
    table_title.pack(side="left", padx=12, pady=8)

    # Table container with modern clean border
    table_container = tk.Frame(table_frame, bg=COLORS['bg_primary'], bd=0, relief="flat")
    table_container.pack(fill="both", expand=True, padx=1, pady=(0, 1))

    global table
    table = ReviewTable(table_container)
    vsb = ttk.Scrollbar(table_container, orient="vertical", command=table.yview)
    hsb = ttk.Scrollbar(table_container, orient="horizontal", command=table.xview)
    table.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

    # Use grid within the container
    table.grid(row=0, column=0, sticky="nsew")
    vsb.grid(row=0, column=1, sticky="ns")
    hsb.grid(row=1, column=0, sticky="ew")
    table_container.grid_rowconfigure(0, weight=1)
    table_container.grid_columnconfigure(0, weight=1)

    # Bottom bar - simplified without redundant tick buttons
    bottom = tk.Frame(win)
    bottom.pack(fill="x", padx=10, pady=(0,10))
    def on_export_comprehensive():
        """Export comprehensive Excel with all details"""
        if not table.rows:
            return messagebox.showwarning("Nothing to export", "Run analysis and review first.")

        out = table.export_comprehensive_xlsx(model_var.get().strip())
        if out:
            messagebox.showinfo("Comprehensive Export",
                f"📊 Comprehensive Excel export complete!\n\n"
                f"✅ All columns included with detailed formatting\n"
                f"📈 Summary sheet with statistics\n"
                f"💾 Saved: {out.name}")

    def on_copy_export_column():
        """Copy export column to clipboard for pasting"""
        if not table.rows:
            return messagebox.showwarning("Nothing to copy", "Run analysis and review first.")

        count = table.copy_export_column()
        messagebox.showinfo("Export Column Copied",
            f"📋 Export column copied to clipboard!\n\n"
            f"✅ {count} features marked with 'X'\n"
            f"📝 Ready to paste into destination file\n"
            f"💡 Paste with Ctrl+V where needed")

    def on_export_csv_legacy():
        """Keep CSV export option for compatibility"""
        if not table.rows:
            return messagebox.showwarning("Nothing to export", "Run analysis and review first.")

        out = table.export_reviewed_with_tick(model_var.get().strip())
        messagebox.showinfo("CSV Export", f"📄 CSV exported: {out.name}")

    def on_export_ticked():
        if not table.rows: return messagebox.showwarning("Nothing to export", "Run and review first.")
        out = table.export_ticked_only(model_var.get().strip())
        messagebox.showinfo("Exported", f"Saved {out.name}")
    def on_export_sequential_mega():
        if not table.rows: return messagebox.showwarning("Nothing to export", "Run Sequential MEGA first.")
        out = table.export_sequential_mega_results(model_var.get().strip())
        messagebox.showinfo("Sequential MEGA Export", f"🚀 Sequential MEGA results exported!\n\n📊 Complete analysis with X-marked selections\n💾 Saved: {out.name}")
    def on_learn_from_results():
        if not table.rows: return messagebox.showwarning("Nothing to learn", "Run analysis and review results first.")
        model = model_var.get().strip()

        # Import learning system
        try:
            from enhanced_dictionary import LearningDictionary
            learning_dict = LearningDictionary()

            # Get current table data for learning
            current_rows = table.current_rows_ui()
            learned_count, negative_count = learning_dict.learn_from_results(current_rows, model)

            stats = learning_dict.get_learning_stats()

            messagebox.showinfo("Learning Complete",
                f"🧠 Knowledge updated successfully!\n\n"
                f"📚 Learned {learned_count} positive examples\n"
                f"🚫 Learned {negative_count} negative examples\n"
                f"📊 Total knowledge: {stats['learned_features']} features\n"
                f"🎯 Covers {stats['covered_codes']} different codes\n\n"
                f"Next analysis will use this learned knowledge!")

        except Exception as e:
            messagebox.showerror("Learning Error", f"Failed to learn from results: {str(e)}")
    # OPTIMIZED EXPORT BUTTONS - Only 2 main options + CSV for legacy
    tk.Button(bottom, text="📊 Export Comprehensive Excel", command=on_export_comprehensive,
             bg="#2563EB", fg="white", font=("Arial", 9, "bold"),
             padx=15, pady=5, relief="flat").pack(side="left", padx=(20,8))

    tk.Button(bottom, text="📋 Copy Export Column", command=on_copy_export_column,
             bg="#059669", fg="white", font=("Arial", 9, "bold"),
             padx=15, pady=5, relief="flat").pack(side="left", padx=8)

    # Legacy CSV option (smaller, less prominent)
    tk.Button(bottom, text="📄 CSV", command=on_export_csv_legacy,
             bg="#6B7280", fg="white", font=("Arial", 8),
             padx=8, pady=3, relief="flat").pack(side="left", padx=(15,8))
    tk.Button(bottom, text="🚀 Export Sequential MEGA", command=on_export_sequential_mega,
              bg="#D13438", fg="white").pack(side="left", padx=6)
    tk.Button(bottom, text="🧠 Learn from Results", command=on_learn_from_results,
              bg="#0078D4", fg="white").pack(side="left", padx=6)

    win.mainloop()

if __name__ == "__main__":
    main()
